#!/usr/bin/env python3
"""
Parameter Validation System for Market Regime Formation System

This module provides comprehensive validation for configuration parameters,
including dependency checking, weight sum validation, and range verification
with specific tolerance levels for the end-to-end testing framework.

Features:
- Parameter dependency validation logic
- Weight sum validation with ±0.001 tolerance
- Range checking with type validation
- Cross-parameter dependency verification
- Real-time validation with detailed error reporting
- Hot-reloading capability support

Author: The Augster
Date: 2025-06-19
Version: 1.0.0
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Tuple, Union
import logging
from pathlib import Path
import json
from datetime import datetime
import numpy as np

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ParameterValidationSystem:
    """Comprehensive parameter validation system for Market Regime Formation System"""

    def __init__(self):
        """Initialize validation system"""
        self.validation_rules = self._load_validation_rules()
        self.dependency_rules = self._load_dependency_rules()
        self.tolerance_levels = self._load_tolerance_levels()
        self.validation_results = {}

        logger.info("Parameter Validation System initialized")

    def _load_validation_rules(self) -> Dict[str, Any]:
        """Load comprehensive validation rules for all parameters"""
        return {
            # Triple Straddle Analysis Rules
            'TripleStraddleAnalysis': {
                'rolling_weight': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True, 'dependency': 'rolling_config'},
                'static_weight': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True, 'dependency': 'rolling_config'},
                'atm_straddle_weight': {'type': 'float', 'min': 0.1, 'max': 0.8, 'required': True, 'dependency': 'straddle_weights'},
                'itm1_straddle_weight': {'type': 'float', 'min': 0.1, 'max': 0.6, 'required': True, 'dependency': 'straddle_weights'},
                'otm1_straddle_weight': {'type': 'float', 'min': 0.1, 'max': 0.6, 'required': True, 'dependency': 'straddle_weights'},
                'correlation_threshold_atm_itm1': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True},
                'correlation_threshold_atm_otm1': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True},
                'correlation_threshold_itm1_otm1': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True},
                'timeframe_3min_weight': {'type': 'float', 'min': 0.0, 'max': 0.5, 'required': True, 'dependency': 'timeframe_weights'},
                'timeframe_5min_weight': {'type': 'float', 'min': 0.0, 'max': 0.5, 'required': True, 'dependency': 'timeframe_weights'},
                'timeframe_10min_weight': {'type': 'float', 'min': 0.0, 'max': 0.5, 'required': True, 'dependency': 'timeframe_weights'},
                'timeframe_15min_weight': {'type': 'float', 'min': 0.0, 'max': 0.5, 'required': True, 'dependency': 'timeframe_weights'},
                'min_confidence_threshold': {'type': 'float', 'min': 0.0, 'max': 1.0, 'required': True},
                'max_processing_time_ms': {'type': 'int', 'min': 1000, 'max': 5000, 'required': True}
            },

            # Greek Sentiment Analysis Rules
            'GreekSentimentAnalysis': {
                'dte_near_expiry_min': {'type': 'int', 'min': 0, 'max': 7, 'required': True, 'dependency': 'dte_ranges'},
                'dte_near_expiry_max': {'type': 'int', 'min': 0, 'max': 14, 'required': True, 'dependency': 'dte_ranges'},
                'dte_medium_expiry_min': {'type': 'int', 'min': 7, 'max': 21, 'required': True, 'dependency': 'dte_ranges'},
                'dte_medium_expiry_max': {'type': 'int', 'min': 14, 'max': 45, 'required': True, 'dependency': 'dte_ranges'},
                'dte_far_expiry_min': {'type': 'int', 'min': 21, 'max': 60, 'required': True, 'dependency': 'dte_ranges'},
                'dte_far_expiry_max': {'type': 'int', 'min': 60, 'max': 180, 'required': True, 'dependency': 'dte_ranges'},
                'delta_weight_near': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'delta_weight_medium': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'delta_weight_far': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'vega_weight_near': {'type': 'float', 'min': 0.3, 'max': 2.5, 'required': True},
                'vega_weight_medium': {'type': 'float', 'min': 0.3, 'max': 2.5, 'required': True},
                'vega_weight_far': {'type': 'float', 'min': 0.3, 'max': 2.5, 'required': True},
                'theta_weight_near': {'type': 'float', 'min': 0.2, 'max': 2.0, 'required': True},
                'theta_weight_medium': {'type': 'float', 'min': 0.2, 'max': 2.0, 'required': True},
                'theta_weight_far': {'type': 'float', 'min': 0.2, 'max': 2.0, 'required': True},
                'gamma_weight_near': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'gamma_weight_medium': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'gamma_weight_far': {'type': 'float', 'min': 0.5, 'max': 2.0, 'required': True},
                'session_baseline_alpha': {'type': 'float', 'min': 0.01, 'max': 0.5, 'required': True},
                'min_oi_threshold': {'type': 'int', 'min': 10, 'max': 1000, 'required': True}
            },

            # Trending OI with PA Rules
            'TrendingOIWithPA': {
                'volume_weighting_factor': {'type': 'float', 'min': 0.1, 'max': 3.0, 'required': True},
                'oi_weighting_factor': {'type': 'float', 'min': 0.1, 'max': 3.0, 'required': True},
                'strike_range_atm_offset': {'type': 'int', 'min': 3, 'max': 15, 'required': True},
                'total_strikes_analyzed': {'type': 'int', 'min': 7, 'max': 25, 'required': True},
                'timeframe_3min_weight': {'type': 'float', 'min': 0.1, 'max': 0.8, 'required': True, 'dependency': 'dual_timeframe_weights'},
                'timeframe_15min_weight': {'type': 'float', 'min': 0.2, 'max': 0.9, 'required': True, 'dependency': 'dual_timeframe_weights'},
                'correlation_threshold_strong': {'type': 'float', 'min': 0.5, 'max': 0.9, 'required': True},
                'correlation_threshold_moderate': {'type': 'float', 'min': 0.3, 'max': 0.7, 'required': True},
                'correlation_threshold_weak': {'type': 'float', 'min': 0.1, 'max': 0.5, 'required': True},
                'price_action_correlation_threshold': {'type': 'float', 'min': 0.3, 'max': 0.9, 'required': True},
                'price_momentum_weight': {'type': 'float', 'min': 0.1, 'max': 0.8, 'required': True, 'dependency': 'momentum_weights'},
                'volume_momentum_weight': {'type': 'float', 'min': 0.1, 'max': 0.8, 'required': True, 'dependency': 'momentum_weights'},
                'oi_momentum_weight': {'type': 'float', 'min': 0.1, 'max': 0.8, 'required': True, 'dependency': 'momentum_weights'},
                'min_volume_threshold': {'type': 'int', 'min': 10, 'max': 1000, 'required': True},
                'min_oi_threshold': {'type': 'int', 'min': 5, 'max': 500, 'required': True}
            },

            # IV Analysis Rules
            'IVAnalysis': {
                'iv_percentile_window': {'type': 'int', 'min': 10, 'max': 50, 'required': True},
                'percentile_threshold_high': {'type': 'int', 'min': 70, 'max': 95, 'required': True},
                'percentile_threshold_low': {'type': 'int', 'min': 5, 'max': 30, 'required': True},
                'skew_threshold_positive': {'type': 'float', 'min': 0.05, 'max': 0.3, 'required': True},
                'skew_threshold_negative': {'type': 'float', 'min': -0.3, 'max': -0.05, 'required': True},
                'surface_smoothing_factor': {'type': 'float', 'min': 0.01, 'max': 0.5, 'required': True},
                'term_structure_contango_threshold': {'type': 'float', 'min': 0.01, 'max': 0.2, 'required': True},
                'term_structure_backwardation_threshold': {'type': 'float', 'min': -0.2, 'max': -0.01, 'required': True},
                'smile_symmetry_threshold': {'type': 'float', 'min': 0.05, 'max': 0.3, 'required': True},
                'smile_steepness_threshold': {'type': 'float', 'min': 0.1, 'max': 0.5, 'required': True},
                'clustering_window': {'type': 'int', 'min': 5, 'max': 20, 'required': True},
                'clustering_threshold': {'type': 'float', 'min': 0.05, 'max': 0.3, 'required': True},
                'iv_weight_in_regime': {'type': 'float', 'min': 0.05, 'max': 0.25, 'required': True},
                'iv_confidence_threshold': {'type': 'float', 'min': 0.4, 'max': 0.9, 'required': True}
            },

            # ATR Technical Rules
            'ATRTechnical': {
                'atr_period': {'type': 'int', 'min': 5, 'max': 30, 'required': True},
                'atr_normalization_period': {'type': 'int', 'min': 10, 'max': 50, 'required': True},
                'breakout_threshold_multiplier': {'type': 'float', 'min': 1.0, 'max': 5.0, 'required': True},
                'breakout_confirmation_periods': {'type': 'int', 'min': 1, 'max': 10, 'required': True},
                'rsi_period': {'type': 'int', 'min': 5, 'max': 30, 'required': True},
                'macd_fast': {'type': 'int', 'min': 5, 'max': 20, 'required': True},
                'macd_slow': {'type': 'int', 'min': 15, 'max': 40, 'required': True},
                'macd_signal': {'type': 'int', 'min': 5, 'max': 15, 'required': True},
                'volume_sma_period': {'type': 'int', 'min': 10, 'max': 50, 'required': True},
                'atr_weight_in_regime': {'type': 'float', 'min': 0.05, 'max': 0.25, 'required': True},
                'atr_confidence_threshold': {'type': 'float', 'min': 0.4, 'max': 0.9, 'required': True}
            },

            # Regime Classification Rules
            'RegimeClassification': {
                'volatility_levels': {'type': 'int', 'min': 2, 'max': 5, 'required': True},
                'directional_levels': {'type': 'int', 'min': 2, 'max': 3, 'required': True},
                'structure_levels': {'type': 'int', 'min': 2, 'max': 3, 'required': True},
                'total_regimes': {'type': 'int', 'min': 8, 'max': 20, 'required': True},
                'component_weight_triple_straddle': {'type': 'float', 'min': 0.20, 'max': 0.50, 'required': True, 'dependency': 'component_weights'},
                'component_weight_greek_sentiment': {'type': 'float', 'min': 0.15, 'max': 0.40, 'required': True, 'dependency': 'component_weights'},
                'component_weight_trending_oi': {'type': 'float', 'min': 0.10, 'max': 0.35, 'required': True, 'dependency': 'component_weights'},
                'component_weight_iv_analysis': {'type': 'float', 'min': 0.05, 'max': 0.25, 'required': True, 'dependency': 'component_weights'},
                'component_weight_atr_technical': {'type': 'float', 'min': 0.05, 'max': 0.25, 'required': True, 'dependency': 'component_weights'},
                'weight_sum_validation': {'type': 'float', 'min': 0.99, 'max': 1.01, 'required': True},
                'weight_tolerance': {'type': 'float', 'min': 0.0001, 'max': 0.01, 'required': True},
                'confidence_threshold_minimum': {'type': 'float', 'min': 0.4, 'max': 0.9, 'required': True},
                'regime_persistence_minimum_minutes': {'type': 'int', 'min': 1, 'max': 10, 'required': True},
                'rapid_switching_threshold': {'type': 'int', 'min': 1, 'max': 5, 'required': True},
                'max_processing_time_ms': {'type': 'int', 'min': 1000, 'max': 5000, 'required': True}
            }
        }

    def _load_dependency_rules(self) -> Dict[str, Any]:
        """Load parameter dependency rules for cross-validation"""
        return {
            # Rolling Configuration Dependencies
            'rolling_config': {
                'description': '100% rolling configuration validation',
                'rules': [
                    {
                        'condition': 'rolling_weight == 1.0',
                        'requirement': 'static_weight == 0.0',
                        'error_message': 'For 100% rolling configuration, static_weight must be 0.0'
                    }
                ]
            },

            # Straddle Weight Dependencies
            'straddle_weights': {
                'description': 'Triple straddle component weight validation',
                'rules': [
                    {
                        'condition': 'sum_validation',
                        'parameters': ['atm_straddle_weight', 'itm1_straddle_weight', 'otm1_straddle_weight'],
                        'target_sum': 1.0,
                        'tolerance': 0.001,
                        'error_message': 'Straddle weights must sum to 1.0 ±0.001'
                    }
                ]
            },

            # Timeframe Weight Dependencies
            'timeframe_weights': {
                'description': 'Multi-timeframe weight validation',
                'rules': [
                    {
                        'condition': 'sum_validation',
                        'parameters': ['timeframe_3min_weight', 'timeframe_5min_weight', 'timeframe_10min_weight', 'timeframe_15min_weight'],
                        'target_sum': 1.0,
                        'tolerance': 0.001,
                        'error_message': 'Timeframe weights must sum to 1.0 ±0.001'
                    }
                ]
            },

            # Dual Timeframe Weight Dependencies (for Trending OI)
            'dual_timeframe_weights': {
                'description': 'Dual timeframe weight validation for Trending OI',
                'rules': [
                    {
                        'condition': 'sum_validation',
                        'parameters': ['timeframe_3min_weight', 'timeframe_15min_weight'],
                        'target_sum': 1.0,
                        'tolerance': 0.001,
                        'error_message': 'Dual timeframe weights must sum to 1.0 ±0.001'
                    }
                ]
            },

            # Momentum Weight Dependencies
            'momentum_weights': {
                'description': 'Momentum component weight validation',
                'rules': [
                    {
                        'condition': 'sum_validation',
                        'parameters': ['price_momentum_weight', 'volume_momentum_weight', 'oi_momentum_weight'],
                        'target_sum': 1.0,
                        'tolerance': 0.001,
                        'error_message': 'Momentum weights must sum to 1.0 ±0.001'
                    }
                ]
            },

            # DTE Range Dependencies
            'dte_ranges': {
                'description': 'DTE range validation for non-overlapping categories',
                'rules': [
                    {
                        'condition': 'range_validation',
                        'parameters': [
                            ('dte_near_expiry_min', 'dte_near_expiry_max'),
                            ('dte_medium_expiry_min', 'dte_medium_expiry_max'),
                            ('dte_far_expiry_min', 'dte_far_expiry_max')
                        ],
                        'requirement': 'non_overlapping_ascending',
                        'error_message': 'DTE ranges must be non-overlapping and ascending'
                    }
                ]
            },

            # Component Weight Dependencies (Main System)
            'component_weights': {
                'description': 'Main component weight validation (35%/25%/20%/10%/10%)',
                'rules': [
                    {
                        'condition': 'sum_validation',
                        'parameters': [
                            'component_weight_triple_straddle',
                            'component_weight_greek_sentiment',
                            'component_weight_trending_oi',
                            'component_weight_iv_analysis',
                            'component_weight_atr_technical'
                        ],
                        'target_sum': 1.0,
                        'tolerance': 0.001,
                        'error_message': 'Component weights must sum to 1.0 ±0.001'
                    }
                ]
            }
        }

    def _load_tolerance_levels(self) -> Dict[str, float]:
        """Load tolerance levels for different validation types"""
        return {
            'mathematical_calculations': 0.001,  # ±0.001 for mathematical accuracy
            'confidence_scores': 0.05,           # ±0.05 for confidence variance
            'weight_sum_validation': 0.001,      # ±0.001 for weight sum validation
            'correlation_thresholds': 0.01,      # ±0.01 for correlation adjustments
            'performance_timing': 100,           # ±100ms for performance timing
            'percentage_values': 0.01            # ±0.01 for percentage-based values
        }

    def validate_excel_configuration(self, excel_path: str) -> Dict[str, Any]:
        """Validate complete Excel configuration file"""
        try:
            logger.info(f"Validating Excel configuration: {excel_path}")

            # Load workbook
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            validation_results = {
                'file_path': excel_path,
                'validation_timestamp': datetime.now().isoformat(),
                'overall_status': 'PENDING',
                'sheet_results': {},
                'dependency_results': {},
                'summary': {
                    'total_parameters': 0,
                    'valid_parameters': 0,
                    'invalid_parameters': 0,
                    'warnings': 0,
                    'errors': 0
                }
            }

            # Validate each sheet
            for sheet_name in workbook.sheetnames:
                if sheet_name == 'Metadata':
                    continue

                if sheet_name in self.validation_rules:
                    sheet_result = self._validate_sheet(workbook[sheet_name], sheet_name)
                    validation_results['sheet_results'][sheet_name] = sheet_result

                    # Update summary
                    validation_results['summary']['total_parameters'] += sheet_result['total_parameters']
                    validation_results['summary']['valid_parameters'] += sheet_result['valid_parameters']
                    validation_results['summary']['invalid_parameters'] += sheet_result['invalid_parameters']
                    validation_results['summary']['warnings'] += len(sheet_result['warnings'])
                    validation_results['summary']['errors'] += len(sheet_result['errors'])

            # Validate dependencies
            dependency_results = self._validate_dependencies(workbook)
            validation_results['dependency_results'] = dependency_results
            validation_results['summary']['errors'] += len(dependency_results['errors'])
            validation_results['summary']['warnings'] += len(dependency_results['warnings'])

            # Determine overall status
            if validation_results['summary']['errors'] == 0:
                validation_results['overall_status'] = 'PASS'
            elif validation_results['summary']['errors'] <= 5:
                validation_results['overall_status'] = 'PARTIAL'
            else:
                validation_results['overall_status'] = 'FAIL'

            workbook.close()

            logger.info(f"Validation complete. Status: {validation_results['overall_status']}")
            return validation_results

        except Exception as e:
            logger.error(f"Error validating Excel configuration: {e}")
            return {
                'file_path': excel_path,
                'validation_timestamp': datetime.now().isoformat(),
                'overall_status': 'ERROR',
                'error_message': str(e)
            }

    def _validate_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Validate individual sheet parameters"""
        sheet_rules = self.validation_rules.get(sheet_name, {})

        result = {
            'sheet_name': sheet_name,
            'total_parameters': 0,
            'valid_parameters': 0,
            'invalid_parameters': 0,
            'errors': [],
            'warnings': [],
            'parameter_results': {}
        }

        # Extract parameters from sheet
        parameters = self._extract_parameters_from_sheet(sheet)
        result['total_parameters'] = len(parameters)

        # Validate each parameter
        for param_name, param_value in parameters.items():
            if param_name in sheet_rules:
                param_result = self._validate_parameter(param_name, param_value, sheet_rules[param_name])
                result['parameter_results'][param_name] = param_result

                if param_result['valid']:
                    result['valid_parameters'] += 1
                else:
                    result['invalid_parameters'] += 1
                    result['errors'].extend(param_result['errors'])

                result['warnings'].extend(param_result['warnings'])
            else:
                result['warnings'].append(f"Parameter '{param_name}' not found in validation rules")

        return result

    def _extract_parameters_from_sheet(self, sheet) -> Dict[str, Any]:
        """Extract parameters and values from Excel sheet"""
        parameters = {}

        # Assume parameter name in column A, value in column B
        for row in range(2, sheet.max_row + 1):
            param_name_cell = sheet.cell(row=row, column=1)
            param_value_cell = sheet.cell(row=row, column=2)

            if param_name_cell.value and param_value_cell.value is not None:
                param_name = str(param_name_cell.value).strip()
                param_value = param_value_cell.value

                # Convert string representations of boolean
                if isinstance(param_value, str):
                    if param_value.upper() == 'TRUE':
                        param_value = True
                    elif param_value.upper() == 'FALSE':
                        param_value = False

                parameters[param_name] = param_value

        return parameters

    def _validate_parameter(self, param_name: str, param_value: Any, rules: Dict[str, Any]) -> Dict[str, Any]:
        """Validate individual parameter against rules"""
        result = {
            'parameter_name': param_name,
            'value': param_value,
            'valid': True,
            'errors': [],
            'warnings': []
        }

        # Type validation
        expected_type = rules.get('type')
        if expected_type:
            if not self._validate_type(param_value, expected_type):
                result['valid'] = False
                result['errors'].append(f"Parameter '{param_name}' expected type {expected_type}, got {type(param_value).__name__}")

        # Range validation for numeric types
        if expected_type in ['int', 'float'] and isinstance(param_value, (int, float)):
            min_val = rules.get('min')
            max_val = rules.get('max')

            if min_val is not None and param_value < min_val:
                result['valid'] = False
                result['errors'].append(f"Parameter '{param_name}' value {param_value} below minimum {min_val}")

            if max_val is not None and param_value > max_val:
                result['valid'] = False
                result['errors'].append(f"Parameter '{param_name}' value {param_value} above maximum {max_val}")

        # Required validation
        if rules.get('required', False) and param_value is None:
            result['valid'] = False
            result['errors'].append(f"Parameter '{param_name}' is required but not provided")

        return result

    def _validate_type(self, value: Any, expected_type: str) -> bool:
        """Validate parameter type"""
        if expected_type == 'int':
            return isinstance(value, int) and not isinstance(value, bool)
        elif expected_type == 'float':
            return isinstance(value, (int, float)) and not isinstance(value, bool)
        elif expected_type == 'bool':
            return isinstance(value, bool)
        elif expected_type == 'str':
            return isinstance(value, str)
        else:
            return True

    def _validate_dependencies(self, workbook) -> Dict[str, Any]:
        """Validate parameter dependencies across sheets"""
        result = {
            'total_dependencies': len(self.dependency_rules),
            'valid_dependencies': 0,
            'invalid_dependencies': 0,
            'errors': [],
            'warnings': [],
            'dependency_results': {}
        }

        # Extract all parameters from all sheets
        all_parameters = {}
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'Metadata':
                sheet_params = self._extract_parameters_from_sheet(workbook[sheet_name])
                all_parameters.update(sheet_params)

        # Validate each dependency rule
        for dependency_name, dependency_rule in self.dependency_rules.items():
            dependency_result = self._validate_dependency_rule(dependency_name, dependency_rule, all_parameters)
            result['dependency_results'][dependency_name] = dependency_result

            if dependency_result['valid']:
                result['valid_dependencies'] += 1
            else:
                result['invalid_dependencies'] += 1
                result['errors'].extend(dependency_result['errors'])

            result['warnings'].extend(dependency_result['warnings'])

        return result

    def _validate_dependency_rule(self, dependency_name: str, dependency_rule: Dict[str, Any], parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Validate specific dependency rule"""
        result = {
            'dependency_name': dependency_name,
            'description': dependency_rule.get('description', ''),
            'valid': True,
            'errors': [],
            'warnings': []
        }

        rules = dependency_rule.get('rules', [])

        for rule in rules:
            condition = rule.get('condition')

            if condition == 'sum_validation':
                # Validate sum of parameters
                param_names = rule.get('parameters', [])
                target_sum = rule.get('target_sum', 1.0)
                tolerance = rule.get('tolerance', 0.001)

                param_values = []
                missing_params = []

                for param_name in param_names:
                    if param_name in parameters:
                        param_values.append(parameters[param_name])
                    else:
                        missing_params.append(param_name)

                if missing_params:
                    result['valid'] = False
                    result['errors'].append(f"Missing parameters for sum validation: {missing_params}")
                else:
                    actual_sum = sum(param_values)
                    if abs(actual_sum - target_sum) > tolerance:
                        result['valid'] = False
                        result['errors'].append(f"{rule.get('error_message', 'Sum validation failed')}: {actual_sum:.6f} vs {target_sum:.6f}")

            elif condition == 'range_validation':
                # Validate range parameters
                param_ranges = rule.get('parameters', [])

                for min_param, max_param in param_ranges:
                    if min_param in parameters and max_param in parameters:
                        min_val = parameters[min_param]
                        max_val = parameters[max_param]

                        if min_val >= max_val:
                            result['valid'] = False
                            result['errors'].append(f"Range validation failed: {min_param}({min_val}) >= {max_param}({max_val})")

            elif condition.startswith('rolling_weight'):
                # Validate rolling configuration
                if 'rolling_weight' in parameters and 'static_weight' in parameters:
                    rolling_weight = parameters['rolling_weight']
                    static_weight = parameters['static_weight']

                    if rolling_weight == 1.0 and static_weight != 0.0:
                        result['valid'] = False
                        result['errors'].append(rule.get('error_message', 'Rolling configuration validation failed'))

        return result

    def generate_validation_report(self, validation_results: Dict[str, Any], output_path: str = None) -> str:
        """Generate comprehensive validation report"""
        if output_path is None:
            output_path = f"parameter_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"

        with open(output_path, 'w') as f:
            f.write("# Parameter Validation Report\n\n")
            f.write(f"**Validation Date:** {validation_results['validation_timestamp']}\n")
            f.write(f"**Configuration File:** {validation_results['file_path']}\n")
            f.write(f"**Overall Status:** {validation_results['overall_status']}\n\n")

            # Summary
            summary = validation_results['summary']
            f.write("## Validation Summary\n\n")
            f.write(f"- **Total Parameters:** {summary['total_parameters']}\n")
            f.write(f"- **Valid Parameters:** {summary['valid_parameters']}\n")
            f.write(f"- **Invalid Parameters:** {summary['invalid_parameters']}\n")
            f.write(f"- **Warnings:** {summary['warnings']}\n")
            f.write(f"- **Errors:** {summary['errors']}\n\n")

            # Sheet Results
            f.write("## Sheet Validation Results\n\n")
            for sheet_name, sheet_result in validation_results.get('sheet_results', {}).items():
                f.write(f"### {sheet_name}\n")
                f.write(f"- **Parameters:** {sheet_result['valid_parameters']}/{sheet_result['total_parameters']} valid\n")
                f.write(f"- **Errors:** {len(sheet_result['errors'])}\n")
                f.write(f"- **Warnings:** {len(sheet_result['warnings'])}\n\n")

                if sheet_result['errors']:
                    f.write("**Errors:**\n")
                    for error in sheet_result['errors']:
                        f.write(f"- {error}\n")
                    f.write("\n")

            # Dependency Results
            dependency_results = validation_results.get('dependency_results', {})
            if dependency_results:
                f.write("## Dependency Validation Results\n\n")
                f.write(f"- **Valid Dependencies:** {dependency_results['valid_dependencies']}/{dependency_results['total_dependencies']}\n")
                f.write(f"- **Errors:** {len(dependency_results['errors'])}\n")
                f.write(f"- **Warnings:** {len(dependency_results['warnings'])}\n\n")

                if dependency_results['errors']:
                    f.write("**Dependency Errors:**\n")
                    for error in dependency_results['errors']:
                        f.write(f"- {error}\n")
                    f.write("\n")

        logger.info(f"Validation report generated: {output_path}")
        return output_path

def validate_all_profiles():
    """Validate all three configuration profiles"""
    validator = ParameterValidationSystem()

    profiles = ['conservative', 'balanced', 'aggressive']
    validation_results = {}

    for profile in profiles:
        excel_path = f"market_regime_unified_config_{profile}.xlsx"
        try:
            result = validator.validate_excel_configuration(excel_path)
            validation_results[profile] = result

            # Generate individual report
            report_path = validator.generate_validation_report(result, f"validation_report_{profile}.md")
            logger.info(f"Validated {profile} profile: {result['overall_status']}")

        except Exception as e:
            logger.error(f"Failed to validate {profile} profile: {e}")
            validation_results[profile] = {'error': str(e)}

    return validation_results

if __name__ == "__main__":
    # Validate all profile configurations
    results = validate_all_profiles()

    print(f"\nParameter Validation Complete!")
    print(f"Validation Results:")
    for profile, result in results.items():
        if 'error' in result:
            print(f"  - {profile.title()}: ERROR - {result['error']}")
        else:
            print(f"  - {profile.title()}: {result['overall_status']}")
            print(f"    Valid: {result['summary']['valid_parameters']}/{result['summary']['total_parameters']} parameters")
            print(f"    Errors: {result['summary']['errors']}, Warnings: {result['summary']['warnings']}")

    print(f"\nValidation Features:")
    print(f"  - Parameter type and range validation")
    print(f"  - Weight sum validation with ±0.001 tolerance")
    print(f"  - Cross-parameter dependency checking")
    print(f"  - 100% rolling configuration validation")
    print(f"  - Comprehensive error reporting")