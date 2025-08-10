"""
Triple Straddle Analysis Excel Configuration System

This module provides comprehensive Excel-based configuration for the Triple Straddle
Analysis system, allowing users to configure all aspects of the analysis including:
- Component weights (ATM, ITM1, OTM1)
- Timeframe weights (3, 5, 10, 15 minutes)
- Technical analysis parameters (EMA, VWAP, Pivot)
- Dynamic weight optimization settings
- Performance tracking parameters
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

logger = logging.getLogger(__name__)

class TripleStraddleExcelConfig:
    """
    Excel Configuration Manager for Triple Straddle Analysis System

    Provides comprehensive Excel-based configuration with the following sheets:
    1. TripleStraddleConfig - Core straddle analysis parameters
    2. WeightOptimization - Dynamic weight adjustment settings
    3. TimeframeSettings - Multi-timeframe analysis configuration
    4. TechnicalAnalysis - EMA/VWAP/Pivot parameters
    5. PerformanceTracking - Historical performance parameters
    6. RegimeThresholds - 18-regime classification thresholds
    """

    def __init__(self, config_path: Optional[str] = None):
        """Initialize Excel configuration manager"""
        self.config_path = config_path or "triple_straddle_config.xlsx"
        self.config_data = {}

        # Default configuration structure
        self.default_config = self._get_default_configuration()

        logger.info(f"Triple Straddle Excel Config initialized: {self.config_path}")

    def generate_excel_template(self, output_path: Optional[str] = None) -> str:
        """
        Generate comprehensive Excel template for Triple Straddle Analysis

        Args:
            output_path: Path for output Excel file

        Returns:
            Path to generated Excel file
        """
        try:
            output_path = output_path or self.config_path

            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create all configuration sheets
            self._create_triple_straddle_config_sheet(wb)
            self._create_weight_optimization_sheet(wb)
            self._create_timeframe_settings_sheet(wb)
            self._create_technical_analysis_sheet(wb)
            self._create_performance_tracking_sheet(wb)
            self._create_regime_thresholds_sheet(wb)

            # Save workbook
            wb.save(output_path)

            logger.info(f"Excel template generated: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel template: {e}")
            raise

    def _create_triple_straddle_config_sheet(self, wb: openpyxl.Workbook):
        """Create Triple Straddle Configuration sheet"""
        ws = wb.create_sheet("TripleStraddleConfig")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Configuration data
        config_data = [
            ["Parameter", "Value", "Description", "Type", "Min", "Max"],
            ["ATMStraddleWeight", 0.50, "Weight for ATM straddle component", "float", 0.1, 0.8],
            ["ITM1StraddleWeight", 0.30, "Weight for ITM1 straddle component", "float", 0.1, 0.6],
            ["OTM1StraddleWeight", 0.20, "Weight for OTM1 straddle component", "float", 0.1, 0.6],
            ["EnableDynamicWeights", "YES", "Enable dynamic weight optimization", "bool", "", ""],
            ["WeightAdjustmentFactor", 0.05, "Factor for weight adjustments", "float", 0.01, 0.20],
            ["MinConfidenceThreshold", 0.60, "Minimum confidence for regime detection", "float", 0.5, 0.9],
            ["MaxVolatilityThreshold", 0.30, "Maximum volatility for normal regime", "float", 0.1, 0.5],
            ["PerformanceWindowSize", 50, "Window size for performance tracking", "int", 20, 200],
            ["OptimizationFrequency", 100, "Frequency of weight optimization (predictions)", "int", 50, 500],
            ["EnableMultiTimeframe", "YES", "Enable multi-timeframe analysis", "bool", "", ""],
            ["EnableATMCEPEAnalysis", "YES", "Enable individual ATM CE/PE analysis", "bool", "", ""],
            ["StraddlePriceCalculation", "SUM", "Method for straddle price calculation", "str", "", ""],
            ["VolumeWeightingEnabled", "YES", "Enable volume weighting in analysis", "bool", "", ""],
            ["IVWeightingEnabled", "YES", "Enable IV weighting in analysis", "bool", "", ""]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(config_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_weight_optimization_sheet(self, wb: openpyxl.Workbook):
        """Create Weight Optimization Configuration sheet"""
        ws = wb.create_sheet("WeightOptimization")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        # Weight optimization data
        optimization_data = [
            ["Parameter", "Value", "Description", "Type", "Min", "Max"],
            ["LearningRate", 0.05, "Learning rate for weight optimization", "float", 0.01, 0.20],
            ["PerformanceWindow", 50, "Window size for performance evaluation", "int", 20, 200],
            ["MinPerformanceThreshold", 0.60, "Minimum performance threshold", "float", 0.5, 0.9],
            ["WeightBoundsMin", 0.05, "Minimum weight bound", "float", 0.01, 0.20],
            ["WeightBoundsMax", 0.60, "Maximum weight bound", "float", 0.40, 0.80],
            ["VolatilityAdjustmentFactor", 0.10, "Factor for volatility-based adjustments", "float", 0.05, 0.30],
            ["TimeOfDayAdjustmentFactor", 0.05, "Factor for time-of-day adjustments", "float", 0.01, 0.15],
            ["OptimizationMethod", "scipy_minimize", "Optimization method to use", "str", "", ""],
            ["ValidationEnabled", "YES", "Enable optimization validation", "bool", "", ""],
            ["RollbackOnFailure", "YES", "Rollback weights on optimization failure", "bool", "", ""],
            ["MaxOptimizationAttempts", 3, "Maximum optimization attempts", "int", 1, 10],
            ["PerformanceImprovementThreshold", 0.02, "Minimum improvement for weight update", "float", 0.01, 0.10],
            ["WeightStabilityFactor", 0.80, "Factor for weight stability", "float", 0.5, 0.95],
            ["EnablePillarOptimization", "YES", "Enable pillar-level optimization", "bool", "", ""],
            ["EnableIndicatorOptimization", "YES", "Enable indicator-level optimization", "bool", "", ""],
            ["EnableComponentOptimization", "YES", "Enable component-level optimization", "bool", "", ""]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(optimization_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_timeframe_settings_sheet(self, wb: openpyxl.Workbook):
        """Create Timeframe Settings Configuration sheet"""
        ws = wb.create_sheet("TimeframeSettings")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")

        # Timeframe settings data
        timeframe_data = [
            ["Timeframe", "Weight", "Description", "Enabled", "Priority"],
            ["3min", 0.10, "Short-term signals and noise filtering", "YES", 4],
            ["5min", 0.30, "Primary analysis timeframe", "YES", 1],
            ["10min", 0.35, "Core trend analysis timeframe", "YES", 2],
            ["15min", 0.25, "Trend confirmation timeframe", "YES", 3],
            ["", "", "", "", ""],
            ["TimeOfDayAdjustments", "", "", "", ""],
            ["MarketOpenHours", "9-10", "Market open adjustment hours", "YES", ""],
            ["MarketOpenWeight3min", 0.20, "3min weight during market open", "YES", ""],
            ["MarketOpenWeight5min", 0.35, "5min weight during market open", "YES", ""],
            ["MarketOpenWeight10min", 0.30, "10min weight during market open", "YES", ""],
            ["MarketOpenWeight15min", 0.15, "15min weight during market open", "YES", ""],
            ["", "", "", "", ""],
            ["MarketCloseHours", "15-16", "Market close adjustment hours", "YES", ""],
            ["MarketCloseWeight3min", 0.05, "3min weight during market close", "YES", ""],
            ["MarketCloseWeight5min", 0.20, "5min weight during market close", "YES", ""],
            ["MarketCloseWeight10min", 0.35, "10min weight during market close", "YES", ""],
            ["MarketCloseWeight15min", 0.40, "15min weight during market close", "YES", ""]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(timeframe_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_technical_analysis_sheet(self, wb: openpyxl.Workbook):
        """Create Technical Analysis Configuration sheet"""
        ws = wb.create_sheet("TechnicalAnalysis")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

        # Technical analysis data
        technical_data = [
            ["Parameter", "Value", "Description", "Type", "Category"],
            ["EMAAnalysisWeight", 0.40, "Weight for EMA analysis", "float", "Technical"],
            ["VWAPAnalysisWeight", 0.35, "Weight for VWAP analysis", "float", "Technical"],
            ["PivotAnalysisWeight", 0.25, "Weight for Pivot analysis", "float", "Technical"],
            ["", "", "", "", ""],
            ["EMA_ShortPeriod", 20, "Short-term EMA period", "int", "EMA"],
            ["EMA_MediumPeriod", 50, "Medium-term EMA period", "int", "EMA"],
            ["EMA_LongPeriod", 100, "Long-term EMA period", "int", "EMA"],
            ["EMA_TrendFilterPeriod", 200, "Trend filter EMA period", "int", "EMA"],
            ["EMA_SlopeThreshold", 0.001, "Minimum slope for trend detection", "float", "EMA"],
            ["", "", "", "", ""],
            ["VWAP_CurrentDayEnabled", "YES", "Enable current day VWAP", "bool", "VWAP"],
            ["VWAP_PreviousDayEnabled", "YES", "Enable previous day VWAP", "bool", "VWAP"],
            ["VWAP_DeviationThreshold1", 0.005, "First deviation threshold (0.5%)", "float", "VWAP"],
            ["VWAP_DeviationThreshold2", 0.020, "Second deviation threshold (2%)", "float", "VWAP"],
            ["VWAP_VolumeWeightingEnabled", "YES", "Enable volume weighting", "bool", "VWAP"],
            ["", "", "", "", ""],
            ["Pivot_CalculationMethod", "STANDARD", "Pivot calculation method", "str", "Pivot"],
            ["Pivot_ResistanceSupport", "YES", "Calculate R1/S1 levels", "bool", "Pivot"],
            ["Pivot_MultipleTimeframes", "YES", "Use multiple timeframes", "bool", "Pivot"],
            ["Pivot_BreakoutThreshold", 0.002, "Breakout threshold (0.2%)", "float", "Pivot"]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(technical_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_performance_tracking_sheet(self, wb: openpyxl.Workbook):
        """Create Performance Tracking Configuration sheet"""
        ws = wb.create_sheet("PerformanceTracking")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Performance tracking data
        performance_data = [
            ["Parameter", "Value", "Description", "Type", "Category"],
            ["PerformanceHistorySize", 1000, "Maximum performance history size", "int", "Tracking"],
            ["RegimeHistorySize", 1000, "Maximum regime history size", "int", "Tracking"],
            ["WeightHistorySize", 100, "Maximum weight history size", "int", "Tracking"],
            ["AccuracyThreshold", 0.70, "Accuracy threshold for good performance", "float", "Metrics"],
            ["PrecisionThreshold", 0.65, "Precision threshold for good performance", "float", "Metrics"],
            ["RecallThreshold", 0.65, "Recall threshold for good performance", "float", "Metrics"],
            ["F1ScoreThreshold", 0.65, "F1 score threshold for good performance", "float", "Metrics"],
            ["ConfidenceThreshold", 0.60, "Confidence threshold for reliable predictions", "float", "Metrics"],
            ["RegimeStabilityThreshold", 0.80, "Regime stability threshold", "float", "Metrics"],
            ["", "", "", "", ""],
            ["EnablePerformanceLogging", "YES", "Enable performance logging", "bool", "Logging"],
            ["LoggingLevel", "INFO", "Logging level", "str", "Logging"],
            ["LoggingFrequency", 10, "Logging frequency (predictions)", "int", "Logging"],
            ["EnableMetricsExport", "YES", "Enable metrics export", "bool", "Export"],
            ["MetricsExportPath", "performance_metrics.csv", "Path for metrics export", "str", "Export"],
            ["ExportFrequency", 100, "Export frequency (predictions)", "int", "Export"]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(performance_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_regime_thresholds_sheet(self, wb: openpyxl.Workbook):
        """Create Regime Thresholds Configuration sheet"""
        ws = wb.create_sheet("RegimeThresholds")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")

        # Regime thresholds data
        regime_data = [
            ["Parameter", "Value", "Description", "Type", "Category"],
            ["StrongBullishThreshold", 0.50, "Threshold for strong bullish regime", "float", "Directional"],
            ["MildBullishThreshold", 0.20, "Threshold for mild bullish regime", "float", "Directional"],
            ["NeutralThreshold", 0.10, "Threshold for neutral regime", "float", "Directional"],
            ["SidewaysThreshold", 0.05, "Threshold for sideways regime", "float", "Directional"],
            ["MildBearishThreshold", -0.20, "Threshold for mild bearish regime", "float", "Directional"],
            ["StrongBearishThreshold", -0.50, "Threshold for strong bearish regime", "float", "Directional"],
            ["", "", "", "", ""],
            ["HighVolatilityThreshold", 0.20, "Threshold for high volatility", "float", "Volatility"],
            ["NormalHighVolatilityThreshold", 0.15, "Threshold for normal-high volatility", "float", "Volatility"],
            ["NormalLowVolatilityThreshold", 0.10, "Threshold for normal-low volatility", "float", "Volatility"],
            ["LowVolatilityThreshold", 0.05, "Threshold for low volatility", "float", "Volatility"],
            ["", "", "", "", ""],
            ["RegimeChangeThreshold", 0.15, "Threshold for regime change detection", "float", "Change"],
            ["RegimeStabilityPeriod", 5, "Periods for regime stability", "int", "Change"],
            ["MinRegimeDuration", 3, "Minimum regime duration (minutes)", "int", "Change"],
            ["MaxRegimeFluctuations", 2, "Maximum fluctuations before regime change", "int", "Change"]
        ]

        # Add data to sheet
        for row_idx, row_data in enumerate(regime_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _get_default_configuration(self) -> Dict[str, Any]:
        """Get default configuration structure"""
        return {
            'triple_straddle': {
                'atm_weight': 0.50,
                'itm1_weight': 0.30,
                'otm1_weight': 0.20,
                'enable_dynamic_weights': True,
                'weight_adjustment_factor': 0.05
            },
            'timeframes': {
                '3min': 0.10,
                '5min': 0.30,
                '10min': 0.35,
                '15min': 0.25
            },
            'technical_analysis': {
                'ema_weight': 0.40,
                'vwap_weight': 0.35,
                'pivot_weight': 0.25,
                'ema_periods': [20, 50, 100, 200]
            },
            'performance_tracking': {
                'history_size': 1000,
                'accuracy_threshold': 0.70,
                'confidence_threshold': 0.60
            }
        }