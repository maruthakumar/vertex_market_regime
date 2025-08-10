#!/usr/bin/env python3
"""
Master Excel Configuration Template Generator
Creates comprehensive configuration template for Market Regime Formation System
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class MasterConfigTemplateGenerator:
    """
    Generates comprehensive Excel configuration template with 9 sheets:
    1. Master Control Panel
    2. Adaptive Timeframe Configuration
    3. Triple Straddle Configuration
    4. Greek Sentiment Configuration
    5. OI Trending Configuration
    6. IV Analysis Configuration
    7. Technical Indicators (ATR, S/R)
    8. ML Model Configuration
    9. Optimizer Settings
    """
    
    def __init__(self):
        self.wb = None
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.sub_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_master_control_panel(self):
        """Create Sheet 1: Master Control Panel"""
        ws = self.wb.active
        ws.title = "Master Control Panel"
        
        # Title
        ws['A1'] = "Market Regime Formation System - Master Control Panel"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Component Weights Section
        ws['A3'] = "Component Weights Configuration"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:E3')
        
        # Component weights data
        component_data = [
            ["Component", "Default Weight", "Min Weight", "Max Weight", "Description"],
            ["Triple Straddle Engine", 0.25, 0.1, 0.4, "Options straddle analysis"],
            ["Greek Sentiment Analyzer", 0.20, 0.1, 0.3, "Options Greeks analysis"],
            ["OI Trending Patterns", 0.20, 0.1, 0.3, "Open Interest patterns"],
            ["IV Analysis Suite", 0.15, 0.05, 0.25, "Implied Volatility analysis"],
            ["ATR Indicators", 0.10, 0.05, 0.2, "Average True Range analysis"],
            ["Support/Resistance", 0.10, 0.05, 0.2, "Technical levels analysis"]
        ]
        
        for row_idx, row_data in enumerate(component_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Timeframe Configuration
        ws['A13'] = "Timeframe Analysis Configuration"
        ws['A13'].font = Font(bold=True, size=14)
        ws.merge_cells('A13:D13')
        
        timeframe_data = [
            ["Timeframe", "Weight", "Enabled", "Description"],
            ["5 minutes", 0.15, "YES", "Ultra short-term analysis"],
            ["15 minutes", 0.25, "YES", "Short-term analysis"],
            ["30 minutes", 0.30, "YES", "Medium-term analysis"],
            ["1 hour", 0.30, "YES", "Longer-term analysis"]
        ]
        
        for row_idx, row_data in enumerate(timeframe_data, start=15):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 15:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # DTE Bucket Configuration
        ws['A21'] = "DTE (Days to Expiry) Bucket Configuration"
        ws['A21'].font = Font(bold=True, size=14)
        ws.merge_cells('A21:E21')
        
        dte_data = [
            ["DTE Range", "Bucket Name", "Weight Multiplier", "ML Model", "Description"],
            ["0-1", "Expiry", 1.5, "XGBoost", "Expiry day special handling"],
            ["2-3", "Very Short", 1.2, "RandomForest", "Very short-term options"],
            ["4-7", "Weekly", 1.0, "RandomForest", "Weekly options"],
            ["8-15", "Bi-Weekly", 0.9, "LightGBM", "Two-week options"],
            ["16-30", "Monthly", 0.8, "NeuralNetwork", "Monthly options"],
            [">30", "Long-term", 0.7, "RandomForest", "Long-term options"]
        ]
        
        for row_idx, row_data in enumerate(dte_data, start=23):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 23:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Regime Definitions
        ws['A31'] = "18-Regime Classification Matrix"
        ws['A31'].font = Font(bold=True, size=14)
        ws.merge_cells('A31:F31')
        
        regime_data = [
            ["Regime ID", "Volatility", "Trend", "Structure", "Name", "Color Code"],
            [0, "Low", "Bullish", "Trending", "Low Vol Bullish Trending", "#00FF00"],
            [1, "Low", "Bullish", "Mean-Rev", "Low Vol Bullish Mean-Rev", "#00CC00"],
            [2, "Low", "Neutral", "Trending", "Low Vol Neutral Trending", "#FFFF00"],
            [3, "Low", "Neutral", "Mean-Rev", "Low Vol Neutral Mean-Rev", "#FFCC00"],
            [4, "Low", "Bearish", "Trending", "Low Vol Bearish Trending", "#FF6600"],
            [5, "Low", "Bearish", "Mean-Rev", "Low Vol Bearish Mean-Rev", "#FF3300"],
            [6, "Medium", "Bullish", "Trending", "Med Vol Bullish Trending", "#00FF66"],
            [7, "Medium", "Bullish", "Mean-Rev", "Med Vol Bullish Mean-Rev", "#00CC66"],
            [8, "Medium", "Neutral", "Trending", "Med Vol Neutral Trending", "#FFFF66"],
            [9, "Medium", "Neutral", "Mean-Rev", "Med Vol Neutral Mean-Rev", "#FFCC66"],
            [10, "Medium", "Bearish", "Trending", "Med Vol Bearish Trending", "#FF6666"],
            [11, "Medium", "Bearish", "Mean-Rev", "Med Vol Bearish Mean-Rev", "#FF3366"],
            [12, "High", "Bullish", "Trending", "High Vol Bullish Trending", "#00FFCC"],
            [13, "High", "Bullish", "Mean-Rev", "High Vol Bullish Mean-Rev", "#00CCCC"],
            [14, "High", "Neutral", "Trending", "High Vol Neutral Trending", "#FFFFCC"],
            [15, "High", "Neutral", "Mean-Rev", "High Vol Neutral Mean-Rev", "#FFCCCC"],
            [16, "High", "Bearish", "Trending", "High Vol Bearish Trending", "#FF66CC"],
            [17, "High", "Bearish", "Mean-Rev", "High Vol Bearish Mean-Rev", "#FF33CC"]
        ]
        
        for row_idx, row_data in enumerate(regime_data, start=33):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 33:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                else:
                    if col_idx == 6:  # Color column
                        cell.fill = PatternFill(start_color=value[1:], end_color=value[1:], fill_type="solid")
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_adaptive_timeframe_config(self):
        """Create Sheet 2: Adaptive Timeframe Configuration"""
        ws = self.wb.create_sheet("Adaptive Timeframe Config")
        
        # Title
        ws['A1'] = "Adaptive Timeframe Configuration - Trading Mode Selection"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Trading Mode Selection
        ws['A3'] = "Trading Mode Selection"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:E3')
        
        mode_data = [
            ["Mode", "Description", "Risk Multiplier", "Transition Threshold", "Stability Window"],
            ["Intraday", "Optimized for minutes to hours holding", 0.8, 0.65, 5],
            ["Positional", "Optimized for hours to days holding", 1.5, 0.85, 30],
            ["Hybrid", "Balanced approach for flexible trading", 1.0, 0.75, 15],
            ["Custom", "User-defined configuration", 1.0, 0.75, 15]
        ]
        
        for row_idx, row_data in enumerate(mode_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # Current Active Mode
        ws['A11'] = "Current Active Mode:"
        ws['A11'].font = Font(bold=True)
        ws['B11'] = "Hybrid"
        ws['B11'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Timeframe Configuration by Mode
        ws['A13'] = "Timeframe Configuration by Trading Mode"
        ws['A13'].font = Font(bold=True, size=14)
        ws.merge_cells('A13:J13')
        
        # Timeframe data header
        header_data = ["Timeframe", "Intraday Enable", "Intraday Weight", 
                      "Positional Enable", "Positional Weight",
                      "Hybrid Enable", "Hybrid Weight",
                      "Custom Enable", "Custom Weight", "Description"]
        
        for col_idx, header in enumerate(header_data, start=1):
            cell = ws.cell(row=15, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Timeframe configuration data
        timeframe_configs = [
            ["3 minutes", "YES", 0.30, "NO", 0.00, "NO", 0.00, "", 0.00, "Ultra short-term scalping"],
            ["5 minutes", "YES", 0.30, "NO", 0.00, "YES", 0.20, "", 0.00, "Short-term entry signals"],
            ["10 minutes", "YES", 0.25, "NO", 0.00, "NO", 0.00, "", 0.00, "Trend confirmation"],
            ["15 minutes", "YES", 0.15, "YES", 0.15, "YES", 0.30, "", 0.00, "Standard intraday"],
            ["30 minutes", "NO", 0.00, "YES", 0.25, "YES", 0.30, "", 0.00, "Medium-term trend"],
            ["1 hour", "NO", 0.00, "YES", 0.35, "YES", 0.20, "", 0.00, "Positional trend"],
            ["4 hours", "NO", 0.00, "YES", 0.25, "NO", 0.00, "", 0.00, "Multi-day context"]
        ]
        
        for row_idx, row_data in enumerate(timeframe_configs, start=16):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Highlight enabled timeframes
                if col_idx in [2, 4, 6, 8] and value == "YES":
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif col_idx in [2, 4, 6, 8] and value == "NO":
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                cell.border = self.border
        
        # Custom Configuration Section
        ws['A25'] = "Custom Mode Configuration"
        ws['A25'].font = Font(bold=True, size=14)
        ws.merge_cells('A25:D25')
        
        custom_data = [
            ["Parameter", "Value", "Valid Range", "Description"],
            ["Mode Name", "My Strategy", "", "Custom mode identifier"],
            ["Description", "Custom configuration", "", "Mode description"],
            ["Risk Multiplier", 1.0, "0.1-3.0", "Risk adjustment factor"],
            ["Transition Threshold", 0.75, "0.5-0.95", "Regime transition sensitivity"],
            ["Stability Window", 15, "1-60", "Minutes for regime stability"]
        ]
        
        for row_idx, row_data in enumerate(custom_data, start=27):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 27:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # Performance Optimization
        ws['A34'] = "Performance-Based Weight Optimization"
        ws['A34'].font = Font(bold=True, size=14)
        ws.merge_cells('A34:E34')
        
        perf_data = [
            ["Timeframe", "Historical Accuracy", "Current Weight", "Optimized Weight", "Status"],
            ["3 minutes", 0.75, 0.30, 0.28, "Active"],
            ["5 minutes", 0.82, 0.30, 0.32, "Active"],
            ["10 minutes", 0.78, 0.25, 0.24, "Active"],
            ["15 minutes", 0.85, 0.15, 0.16, "Active"],
            ["30 minutes", 0.80, 0.00, 0.00, "Disabled"],
            ["1 hour", 0.77, 0.00, 0.00, "Disabled"]
        ]
        
        for row_idx, row_data in enumerate(perf_data, start=36):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 36:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # Mode Recommendation
        ws['A44'] = "Automatic Mode Recommendation"
        ws['A44'].font = Font(bold=True, size=14)
        ws.merge_cells('A44:D44')
        
        recommend_data = [
            ["Strategy Parameter", "Value", "Recommended Mode", "Confidence"],
            ["Average Holding Period", "45 minutes", "Intraday", "High"],
            ["Trade Frequency", "High", "Intraday", "High"],
            ["Risk Tolerance", "Medium", "Hybrid", "Medium"],
            ["Market Conditions", "Normal Vol", "Hybrid", "Medium"]
        ]
        
        for row_idx, row_data in enumerate(recommend_data, start=46):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 46:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_triple_straddle_config(self):
        """Create Sheet 3: Triple Straddle Configuration"""
        ws = self.wb.create_sheet("Triple Straddle Config")
        
        # Title
        ws['A1'] = "Triple Straddle Engine Configuration"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Configuration parameters
        config_data = [
            ["Parameter", "Default Value", "Min Value", "Max Value", "Type", "Description"],
            ["ATM Strike Selection", "Synthetic Future", "", "", "Method", "ATM calculation method"],
            ["Strike Range", 3, 1, 5, "Integer", "Number of strikes around ATM"],
            ["Delta Threshold", 0.5, 0.4, 0.6, "Float", "Delta threshold for ATM"],
            ["Premium Weight", 0.7, 0.5, 0.9, "Float", "Weight for premium in calculation"],
            ["Skew Weight", 0.3, 0.1, 0.5, "Float", "Weight for skew analysis"],
            ["Volume Filter", 100, 50, 500, "Integer", "Minimum volume threshold"],
            ["OI Filter", 100, 50, 1000, "Integer", "Minimum OI threshold"],
            ["Update Frequency", "1m", "", "", "String", "Calculation update frequency"],
            ["Decay Factor", 0.95, 0.9, 0.99, "Float", "Exponential decay for history"],
            ["Lookback Period", 20, 10, 50, "Integer", "Historical lookback in minutes"]
        ]
        
        for row_idx, row_data in enumerate(config_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Straddle Value Thresholds
        ws['A15'] = "Straddle Value Interpretation Thresholds"
        ws['A15'].font = Font(bold=True, size=14)
        ws.merge_cells('A15:D15')
        
        threshold_data = [
            ["Regime", "Min Value", "Max Value", "Interpretation"],
            ["Very Bullish", 2.0, 999, "Strong bullish sentiment"],
            ["Bullish", 1.0, 2.0, "Moderate bullish sentiment"],
            ["Neutral", -1.0, 1.0, "Balanced market sentiment"],
            ["Bearish", -2.0, -1.0, "Moderate bearish sentiment"],
            ["Very Bearish", -999, -2.0, "Strong bearish sentiment"]
        ]
        
        for row_idx, row_data in enumerate(threshold_data, start=17):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 17:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_greek_sentiment_config(self):
        """Create Sheet 4: Greek Sentiment Configuration"""
        ws = self.wb.create_sheet("Greek Sentiment Config")
        
        # Title
        ws['A1'] = "Greek Sentiment Analyzer Configuration"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Greek weights configuration
        greek_data = [
            ["Greek", "Default Weight", "Min Weight", "Max Weight", "Calculation Method", "Description"],
            ["Delta", 0.30, 0.2, 0.4, "Net Delta", "Directional exposure"],
            ["Gamma", 0.25, 0.15, 0.35, "Net Gamma", "Rate of delta change"],
            ["Vega", 0.20, 0.1, 0.3, "Net Vega", "Volatility sensitivity"],
            ["Theta", 0.15, 0.1, 0.25, "Net Theta", "Time decay"],
            ["Rho", 0.10, 0.05, 0.15, "Net Rho", "Interest rate sensitivity"]
        ]
        
        for row_idx, row_data in enumerate(greek_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Greek thresholds
        ws['A10'] = "Greek Sentiment Thresholds"
        ws['A10'].font = Font(bold=True, size=14)
        ws.merge_cells('A10:E10')
        
        threshold_data = [
            ["Sentiment Level", "Min Score", "Max Score", "Market Interpretation", "Action"],
            ["Extremely Bullish", 0.7, 1.0, "Very positive Greek flow", "Strong buy signal"],
            ["Bullish", 0.3, 0.7, "Positive Greek flow", "Buy signal"],
            ["Neutral", -0.3, 0.3, "Balanced Greek flow", "Hold position"],
            ["Bearish", -0.7, -0.3, "Negative Greek flow", "Sell signal"],
            ["Extremely Bearish", -1.0, -0.7, "Very negative Greek flow", "Strong sell signal"]
        ]
        
        for row_idx, row_data in enumerate(threshold_data, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Greek calculation parameters
        ws['A19'] = "Greek Calculation Parameters"
        ws['A19'].font = Font(bold=True, size=14)
        ws.merge_cells('A19:D19')
        
        calc_data = [
            ["Parameter", "Default Value", "Range", "Description"],
            ["Risk-free Rate", 0.05, "0.0-0.1", "Annual risk-free rate"],
            ["Dividend Yield", 0.02, "0.0-0.05", "Annual dividend yield"],
            ["Greek Smoothing", 0.8, "0.5-0.95", "Exponential smoothing factor"],
            ["Outlier Threshold", 3.0, "2.0-4.0", "Standard deviations for outlier"],
            ["Min Strike Count", 5, "3-10", "Minimum strikes for calculation"]
        ]
        
        for row_idx, row_data in enumerate(calc_data, start=21):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 21:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_oi_trending_config(self):
        """Create Sheet 5: OI Trending Configuration"""
        ws = self.wb.create_sheet("OI Trending Config")
        
        # Title
        ws['A1'] = "Open Interest Trending Patterns Configuration"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # OI Pattern Detection
        pattern_data = [
            ["Pattern Type", "Detection Method", "Threshold", "Weight", "Timeframe", "Description"],
            ["OI Buildup", "Percentage Change", 20, 0.25, "5m", "Rapid OI increase"],
            ["OI Unwinding", "Percentage Change", -20, 0.25, "5m", "Rapid OI decrease"],
            ["PCR Shift", "Ratio Change", 0.2, 0.20, "15m", "Put-Call Ratio changes"],
            ["Strike Migration", "Volume Weighted", 1000, 0.15, "30m", "OI moving between strikes"],
            ["Max Pain Shift", "Price Distance", 0.5, 0.15, "1h", "Max pain level changes"]
        ]
        
        for row_idx, row_data in enumerate(pattern_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # OI Analysis Parameters
        ws['A10'] = "OI Analysis Parameters"
        ws['A10'].font = Font(bold=True, size=14)
        ws.merge_cells('A10:D10')
        
        param_data = [
            ["Parameter", "Default Value", "Range", "Description"],
            ["Min OI Threshold", 1000, "500-5000", "Minimum OI for analysis"],
            ["OI Change Window", 5, "3-15", "Minutes for change calculation"],
            ["Strike Range", 10, "5-20", "Number of strikes to analyze"],
            ["Volume Confirmation", "YES", "YES/NO", "Require volume confirmation"],
            ["Smoothing Factor", 0.7, "0.5-0.9", "EMA smoothing for OI changes"],
            ["Outlier Filter", 3.0, "2.0-4.0", "Standard deviations for outlier"]
        ]
        
        for row_idx, row_data in enumerate(param_data, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # OI Interpretation Rules
        ws['A20'] = "OI Interpretation Rules"
        ws['A20'].font = Font(bold=True, size=14)
        ws.merge_cells('A20:E20')
        
        rules_data = [
            ["OI Change", "Price Change", "Interpretation", "Market View", "Confidence"],
            ["Increase", "Increase", "Long Buildup", "Bullish", "High"],
            ["Increase", "Decrease", "Short Buildup", "Bearish", "High"],
            ["Decrease", "Increase", "Short Covering", "Bullish", "Medium"],
            ["Decrease", "Decrease", "Long Unwinding", "Bearish", "Medium"],
            ["Stable", "Any", "Consolidation", "Neutral", "Low"]
        ]
        
        for row_idx, row_data in enumerate(rules_data, start=22):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 22:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_iv_analysis_config(self):
        """Create Sheet 6: IV Analysis Configuration"""
        ws = self.wb.create_sheet("IV Analysis Config")
        
        # Title
        ws['A1'] = "Implied Volatility Analysis Suite Configuration"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # IV Components
        iv_data = [
            ["IV Component", "Weight", "Calculation Method", "Update Freq", "Lookback", "Description"],
            ["IV Rank", 0.25, "Percentile", "1m", "30d", "IV percentile ranking"],
            ["IV Percentile", 0.20, "Historical", "5m", "252d", "Long-term IV positioning"],
            ["IV Skew", 0.20, "25D-ATM", "1m", "5d", "Put-Call IV difference"],
            ["Term Structure", 0.15, "Near-Far", "15m", "1d", "IV term structure slope"],
            ["IV Momentum", 0.10, "Rate of Change", "5m", "1h", "IV directional momentum"],
            ["IV Mean Reversion", 0.10, "Z-Score", "30m", "20d", "Mean reversion signal"]
        ]
        
        for row_idx, row_data in enumerate(iv_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # IV Regime Thresholds
        ws['A11'] = "IV Regime Classification Thresholds"
        ws['A11'].font = Font(bold=True, size=14)
        ws.merge_cells('A11:E11')
        
        regime_data = [
            ["IV Regime", "IV Rank Min", "IV Rank Max", "Market Condition", "Trading Bias"],
            ["Ultra Low IV", 0, 20, "Extreme complacency", "Long volatility"],
            ["Low IV", 20, 40, "Below average volatility", "Neutral to long vol"],
            ["Normal IV", 40, 60, "Average volatility", "Balanced approach"],
            ["Elevated IV", 60, 80, "Above average volatility", "Neutral to short vol"],
            ["High IV", 80, 100, "Extreme fear/uncertainty", "Short volatility"]
        ]
        
        for row_idx, row_data in enumerate(regime_data, start=13):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 13:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # IV Analysis Parameters
        ws['A20'] = "IV Analysis Parameters"
        ws['A20'].font = Font(bold=True, size=14)
        ws.merge_cells('A20:D20')
        
        param_data = [
            ["Parameter", "Default Value", "Range", "Description"],
            ["IV Smoothing", "EMA", "SMA/EMA/WMA", "Smoothing method for IV"],
            ["Smoothing Period", 20, "10-50", "Smoothing period in minutes"],
            ["Outlier Threshold", 4.0, "3.0-5.0", "Std dev for IV outliers"],
            ["Min Strikes", 5, "3-10", "Minimum strikes for skew"],
            ["Skew Threshold", 5.0, "3.0-10.0", "Significant skew percentage"],
            ["Term Structure Days", "7,30,60", "", "Days for term structure"]
        ]
        
        for row_idx, row_data in enumerate(param_data, start=22):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 22:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_technical_indicators_config(self):
        """Create Sheet 7: Technical Indicators Configuration"""
        ws = self.wb.create_sheet("Technical Indicators")
        
        # Title
        ws['A1'] = "Technical Indicators Configuration (ATR & Support/Resistance)"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # ATR Configuration
        ws['A3'] = "ATR (Average True Range) Configuration"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:E3')
        
        atr_data = [
            ["Parameter", "Default Value", "Range", "Timeframe", "Description"],
            ["ATR Period", 14, "10-20", "5m", "Number of periods for ATR"],
            ["ATR Multiplier", 2.0, "1.5-3.0", "All", "Multiplier for bands"],
            ["Smoothing Method", "EMA", "SMA/EMA/WMA", "All", "ATR smoothing method"],
            ["Volatility Threshold Low", 0.5, "0.3-0.7", "All", "Low volatility threshold"],
            ["Volatility Threshold High", 1.5, "1.2-2.0", "All", "High volatility threshold"]
        ]
        
        for row_idx, row_data in enumerate(atr_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Support/Resistance Configuration
        ws['A12'] = "Support/Resistance Level Detection"
        ws['A12'].font = Font(bold=True, size=14)
        ws.merge_cells('A12:E12')
        
        sr_data = [
            ["Method", "Weight", "Parameters", "Timeframe", "Description"],
            ["Pivot Points", 0.25, "Standard/Fibonacci", "Daily", "Classical pivot calculations"],
            ["Volume Profile", 0.25, "VWAP/POC/VAH/VAL", "Intraday", "Volume-based levels"],
            ["Historical Levels", 0.20, "Swing High/Low", "Multi-TF", "Previous swing points"],
            ["Round Numbers", 0.15, "100/500/1000 intervals", "All", "Psychological levels"],
            ["Dynamic Levels", 0.15, "MA/Bollinger/Keltner", "Multi-TF", "Moving average based"]
        ]
        
        for row_idx, row_data in enumerate(sr_data, start=14):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 14:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Level Strength Parameters
        ws['A21'] = "Support/Resistance Strength Parameters"
        ws['A21'].font = Font(bold=True, size=14)
        ws.merge_cells('A21:D21')
        
        strength_data = [
            ["Parameter", "Default Value", "Range", "Description"],
            ["Touch Count", 3, "2-5", "Min touches for valid level"],
            ["Level Tolerance", 0.2, "0.1-0.5", "Percentage tolerance band"],
            ["Volume Confirmation", "YES", "YES/NO", "Require volume at level"],
            ["Time Decay Factor", 0.95, "0.9-0.99", "Decay for older levels"],
            ["Max Active Levels", 5, "3-10", "Maximum S/R levels to track"]
        ]
        
        for row_idx, row_data in enumerate(strength_data, start=23):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 23:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_ml_model_config(self):
        """Create Sheet 8: ML Model Configuration"""
        ws = self.wb.create_sheet("ML Model Config")
        
        # Title
        ws['A1'] = "Machine Learning Model Configuration"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Model Selection by DTE
        ws['A3'] = "ML Model Selection by DTE Bucket"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:F3')
        
        model_data = [
            ["DTE Range", "Primary Model", "Secondary Model", "Ensemble Weight", "Update Freq", "Features"],
            ["0-1", "XGBoost", "LightGBM", "0.7/0.3", "1m", "50"],
            ["2-3", "RandomForest", "XGBoost", "0.6/0.4", "5m", "45"],
            ["4-7", "RandomForest", "NeuralNet", "0.5/0.5", "5m", "40"],
            ["8-15", "LightGBM", "RandomForest", "0.6/0.4", "15m", "35"],
            ["16-30", "NeuralNetwork", "LightGBM", "0.5/0.5", "30m", "30"],
            [">30", "RandomForest", "XGBoost", "0.7/0.3", "1h", "25"]
        ]
        
        for row_idx, row_data in enumerate(model_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Model Parameters
        ws['A13'] = "Model Hyperparameters"
        ws['A13'].font = Font(bold=True, size=14)
        ws.merge_cells('A13:E13')
        
        param_data = [
            ["Model", "Key Parameter", "Default Value", "Range", "Description"],
            ["RandomForest", "n_estimators", 100, "50-200", "Number of trees"],
            ["RandomForest", "max_depth", 10, "5-20", "Maximum tree depth"],
            ["XGBoost", "learning_rate", 0.1, "0.01-0.3", "Learning rate"],
            ["XGBoost", "n_estimators", 100, "50-200", "Number of boosting rounds"],
            ["LightGBM", "num_leaves", 31, "20-50", "Number of leaves"],
            ["LightGBM", "learning_rate", 0.1, "0.01-0.3", "Learning rate"],
            ["NeuralNetwork", "hidden_layers", "[64,32,16]", "", "Hidden layer sizes"],
            ["NeuralNetwork", "learning_rate", 0.001, "0.0001-0.01", "Learning rate"]
        ]
        
        for row_idx, row_data in enumerate(param_data, start=15):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 15:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Feature Engineering
        ws['A25'] = "Feature Engineering Configuration"
        ws['A25'].font = Font(bold=True, size=14)
        ws.merge_cells('A25:D25')
        
        feature_data = [
            ["Feature Category", "Features", "Importance", "Engineering Method"],
            ["Price Features", "OHLC, Returns, MA", "High", "Rolling windows"],
            ["Volume Features", "Volume, OBV, VWAP", "Medium", "Cumulative sums"],
            ["Option Features", "IV, Greeks, OI", "High", "Normalized values"],
            ["Technical Features", "RSI, MACD, BB", "Medium", "Standard calculations"],
            ["Market Features", "VIX, Breadth, Sentiment", "Medium", "External data"],
            ["Time Features", "Hour, Day, Month", "Low", "Cyclical encoding"]
        ]
        
        for row_idx, row_data in enumerate(feature_data, start=27):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 27:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def create_optimizer_settings(self):
        """Create Sheet 9: Optimizer Settings"""
        ws = self.wb.create_sheet("Optimizer Settings")
        
        # Title
        ws['A1'] = "System Optimizer Settings"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Optimization Algorithms
        ws['A3'] = "Optimization Algorithm Configuration"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:E3')
        
        algo_data = [
            ["Optimizer", "Enabled", "Update Frequency", "Parameters", "Description"],
            ["Adaptive Rolling Window", "YES", "100 predictions", "window_range=[1,30]", "Dynamic window selection"],
            ["Dynamic Regime Boundary", "YES", "1000 classifications", "max_shift=0.1", "Adaptive regime boundaries"],
            ["Component Weight Optimizer", "YES", "Daily", "method=genetic", "Optimize component weights"],
            ["Feature Selection", "YES", "Weekly", "method=recursive", "Select optimal features"],
            ["Hyperparameter Tuning", "YES", "Weekly", "method=bayesian", "Tune model parameters"]
        ]
        
        for row_idx, row_data in enumerate(algo_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Performance Metrics
        ws['A12'] = "Performance Monitoring Configuration"
        ws['A12'].font = Font(bold=True, size=14)
        ws.merge_cells('A12:D12')
        
        metric_data = [
            ["Metric", "Target Value", "Alert Threshold", "Description"],
            ["Regime Accuracy", 0.85, 0.75, "Correct regime classification rate"],
            ["Transition Accuracy", 0.80, 0.70, "Regime transition prediction accuracy"],
            ["Component Agreement", 0.75, 0.65, "Agreement between components"],
            ["Processing Latency", "100ms", "200ms", "Maximum processing time"],
            ["Memory Usage", "4GB", "6GB", "Maximum memory consumption"],
            ["API Response Time", "50ms", "100ms", "API endpoint response time"]
        ]
        
        for row_idx, row_data in enumerate(metric_data, start=14):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 14:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # System Parameters
        ws['A22'] = "System-Wide Parameters"
        ws['A22'].font = Font(bold=True, size=14)
        ws.merge_cells('A22:D22')
        
        system_data = [
            ["Parameter", "Default Value", "Range", "Description"],
            ["Cache TTL", 300, "60-600", "Cache time-to-live in seconds"],
            ["Max Workers", 16, "4-32", "Maximum parallel workers"],
            ["Batch Size", 1000, "100-5000", "Processing batch size"],
            ["Log Level", "INFO", "DEBUG/INFO/WARN", "Logging verbosity"],
            ["Data Retention", 30, "7-90", "Days to retain historical data"],
            ["Backup Frequency", "Daily", "Hourly/Daily/Weekly", "Configuration backup schedule"]
        ]
        
        for row_idx, row_data in enumerate(system_data, start=24):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 24:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def save_template(self, filename="market_regime_config_template.xlsx"):
        """Save the Excel template"""
        # Create workbook
        self.wb = openpyxl.Workbook()
        
        # Create all sheets
        self.create_master_control_panel()
        self.create_adaptive_timeframe_config()
        self.create_triple_straddle_config()
        self.create_greek_sentiment_config()
        self.create_oi_trending_config()
        self.create_iv_analysis_config()
        self.create_technical_indicators_config()
        self.create_ml_model_config()
        self.create_optimizer_settings()
        
        # Save the workbook
        self.wb.save(filename)
        logger.info(f"Excel configuration template saved as: {filename}")
        
        return filename

# Create and save the template
if __name__ == "__main__":
    generator = MasterConfigTemplateGenerator()
    output_file = generator.save_template("/srv/samba/shared/bt/backtester_stable/BTRUN/backtester_v2/market_regime/market_regime_config_template.xlsx")
    print(f"✓ Master Excel Configuration Template created: {output_file}")
    print("\nTemplate includes 9 comprehensive sheets:")
    print("1. Master Control Panel - Component weights, timeframes, DTE buckets, regime definitions")
    print("2. Adaptive Timeframe Configuration - Trading mode selection, dynamic timeframes, performance optimization")
    print("3. Triple Straddle Configuration - ATM selection, thresholds, parameters")
    print("4. Greek Sentiment Configuration - Greek weights, thresholds, calculations")
    print("5. OI Trending Configuration - Pattern detection, analysis parameters, rules")
    print("6. IV Analysis Configuration - IV components, regime thresholds, parameters")
    print("7. Technical Indicators - ATR configuration, Support/Resistance detection")
    print("8. ML Model Configuration - Model selection by DTE, hyperparameters, features")
    print("9. Optimizer Settings - Algorithm configuration, performance metrics, system parameters")