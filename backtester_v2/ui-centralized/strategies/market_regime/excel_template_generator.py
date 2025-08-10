"""
Market Regime Excel Template Generator
=====================================

This module creates comprehensive Excel configuration templates for the Market Regime Detection System
with support for both 8-regime and 18-regime modes, including all indicator configurations,
strike analysis parameters, and dynamic weighting settings.

Features:
- Comprehensive Excel template generation
- Multiple template types (8-regime, 18-regime, demo)
- Indicator configuration with weights and parameters
- Strike analysis configuration (ATM, ITM1, OTM1, Combined)
- Dynamic weighting parameters
- Multi-timeframe support
- DTE-specific adaptations
- Validation and schema compliance

Author: Market Regime Integration Team
Date: 2025-06-15
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Import enhanced parameter system
try:
    from .archive_enhanced_modules_do_not_use.enhanced_indicator_parameters import EnhancedIndicatorParameters, UserLevel
except ImportError:
    # Handle relative import when running as script
    import sys
    current_dir = Path(__file__).parent
    sys.path.insert(0, str(current_dir))
    from enhanced_indicator_parameters import EnhancedIndicatorParameters, UserLevel

logger = logging.getLogger(__name__)

class MarketRegimeTemplateGenerator:
    """Generate comprehensive Excel templates for Market Regime configuration"""
    
    def __init__(self, output_dir: str = "/srv/samba/shared/bt/backtester_stable/BTRUN/input_sheets"):
        """
        Initialize template generator

        Args:
            output_dir (str): Directory to save generated templates
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        # Initialize enhanced parameter system
        self.enhanced_params = EnhancedIndicatorParameters()

        # Template configurations
        self.indicator_configs = self._get_indicator_configurations()
        self.strike_configs = self._get_strike_configurations()
        self.regime_configs = self._get_regime_configurations()
        self.dynamic_weight_configs = self._get_dynamic_weight_configurations()

        logger.info(f"✅ MarketRegimeTemplateGenerator initialized, output: {self.output_dir}")
        logger.info(f"📊 Enhanced parameters loaded: {len(self.enhanced_params.get_all_indicators())} indicators")
    
    def _get_indicator_configurations(self) -> Dict[str, Any]:
        """Get comprehensive indicator configurations"""
        return {
            'Greek_Sentiment': {
                'enabled': True,
                'base_weight': 1.0,
                'min_weight': 0.05,
                'max_weight': 0.50,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "3,5,10,15",
                'description': "Greek sentiment analysis with Delta, Gamma, Theta, Vega weighting"
            },
            'Trending_OI_PA': {
                'enabled': True,
                'base_weight': 0.9,
                'min_weight': 0.05,
                'max_weight': 0.45,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "Trending Open Interest with Price Action analysis"
            },
            'EMA_ATM': {
                'enabled': True,
                'base_weight': 0.8,
                'min_weight': 0.05,
                'max_weight': 0.40,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "3,5,10,15",
                'description': "EMA analysis for ATM strikes"
            },
            'EMA_OTM1': {
                'enabled': True,
                'base_weight': 0.7,
                'min_weight': 0.05,
                'max_weight': 0.35,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "EMA analysis for OTM1 strikes (+50/100 points)"
            },
            'EMA_ITM1': {
                'enabled': True,
                'base_weight': 0.7,
                'min_weight': 0.05,
                'max_weight': 0.35,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "EMA analysis for ITM1 strikes (-50/100 points)"
            },
            'EMA_Combined': {
                'enabled': True,
                'base_weight': 0.9,
                'min_weight': 0.05,
                'max_weight': 0.45,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "Combined EMA analysis (ATM+ITM1+OTM1 weighted)"
            },
            'VWAP_ATM': {
                'enabled': True,
                'base_weight': 0.8,
                'min_weight': 0.05,
                'max_weight': 0.40,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "3,5,10,15",
                'description': "VWAP analysis for ATM strikes"
            },
            'VWAP_OTM1': {
                'enabled': True,
                'base_weight': 0.7,
                'min_weight': 0.05,
                'max_weight': 0.35,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "VWAP analysis for OTM1 strikes"
            },
            'VWAP_ITM1': {
                'enabled': True,
                'base_weight': 0.7,
                'min_weight': 0.05,
                'max_weight': 0.35,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "VWAP analysis for ITM1 strikes"
            },
            'VWAP_Combined': {
                'enabled': True,
                'base_weight': 0.9,
                'min_weight': 0.05,
                'max_weight': 0.45,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "Combined VWAP analysis (ATM+ITM1+OTM1 weighted)"
            },
            'IV_Skew': {
                'enabled': True,
                'base_weight': 0.6,
                'min_weight': 0.05,
                'max_weight': 0.30,
                'performance_tracking': True,
                'dte_adaptation': False,
                'timeframe_support': "10,15",
                'description': "Implied Volatility Skew analysis"
            },
            'ATR_Indicators': {
                'enabled': True,
                'base_weight': 0.5,
                'min_weight': 0.05,
                'max_weight': 0.25,
                'performance_tracking': True,
                'dte_adaptation': False,
                'timeframe_support': "15",
                'description': "Average True Range indicators"
            },
            'Premium_Indicators': {
                'enabled': True,
                'base_weight': 0.4,
                'min_weight': 0.05,
                'max_weight': 0.20,
                'performance_tracking': True,
                'dte_adaptation': True,
                'timeframe_support': "5,10,15",
                'description': "Premium decay and time value indicators"
            }
        }
    
    def _get_strike_configurations(self) -> Dict[str, Any]:
        """Get strike configuration parameters"""
        return {
            'ATM_Method': {
                'value': 'CLOSEST_TO_UNDERLYING',
                'description': 'Method for ATM strike identification',
                'options': 'CLOSEST_TO_UNDERLYING, ROUND_STRIKE, WEIGHTED_AVERAGE'
            },
            'OTM1_Offset_DTE_0_3': {
                'value': 50,
                'description': 'Points offset for OTM1 strikes when DTE 0-3 days',
                'range': '25-100'
            },
            'OTM1_Offset_DTE_4_Plus': {
                'value': 100,
                'description': 'Points offset for OTM1 strikes when DTE 4+ days',
                'range': '50-200'
            },
            'ITM1_Offset_DTE_0_3': {
                'value': -50,
                'description': 'Points offset for ITM1 strikes when DTE 0-3 days',
                'range': '-100 to -25'
            },
            'ITM1_Offset_DTE_4_Plus': {
                'value': -100,
                'description': 'Points offset for ITM1 strikes when DTE 4+ days',
                'range': '-200 to -50'
            },
            'Strike_Weighting_Method': {
                'value': 'EQUAL_WEIGHTED',
                'description': 'Weighting method for ATM/ITM1/OTM1 combination',
                'options': 'EQUAL_WEIGHTED, ATM_HEAVY, VOLATILITY_WEIGHTED'
            },
            'Combined_Analysis_Enabled': {
                'value': True,
                'description': 'Enable combined strike analysis',
                'options': 'TRUE, FALSE'
            },
            'Multi_Timeframe_Enabled': {
                'value': True,
                'description': 'Enable multi-timeframe strike analysis',
                'options': 'TRUE, FALSE'
            }
        }
    
    def _get_regime_configurations(self) -> Dict[str, Dict[str, Any]]:
        """Get regime formation configurations for both 8 and 18 regime modes"""
        return {
            '8_REGIME': {
                'Strong_Bullish': {'min': 0.7, 'max': 1.0, 'confidence_min': 0.8},
                'Mild_Bullish': {'min': 0.3, 'max': 0.7, 'confidence_min': 0.6},
                'Neutral': {'min': -0.2, 'max': 0.3, 'confidence_min': 0.5},
                'Mild_Bearish': {'min': -0.7, 'max': -0.2, 'confidence_min': 0.6},
                'Strong_Bearish': {'min': -1.0, 'max': -0.7, 'confidence_min': 0.8},
                'High_Volatile': {'atr_min': 0.8, 'confidence_min': 0.7},
                'Low_Volatile': {'atr_max': 0.2, 'confidence_min': 0.6},
                'Sideways': {'directional_max': 0.2, 'volatility_max': 0.3, 'confidence_min': 0.5}
            },
            '18_REGIME': {
                'Strong_Bullish_High_Vol': {'min': 0.7, 'max': 1.0, 'vol_min': 0.6, 'confidence_min': 0.8},
                'Strong_Bullish_Med_Vol': {'min': 0.7, 'max': 1.0, 'vol_min': 0.3, 'vol_max': 0.6, 'confidence_min': 0.8},
                'Strong_Bullish_Low_Vol': {'min': 0.7, 'max': 1.0, 'vol_max': 0.3, 'confidence_min': 0.8},
                'Mild_Bullish_High_Vol': {'min': 0.3, 'max': 0.7, 'vol_min': 0.6, 'confidence_min': 0.6},
                'Mild_Bullish_Med_Vol': {'min': 0.3, 'max': 0.7, 'vol_min': 0.3, 'vol_max': 0.6, 'confidence_min': 0.6},
                'Mild_Bullish_Low_Vol': {'min': 0.3, 'max': 0.7, 'vol_max': 0.3, 'confidence_min': 0.6},
                'Neutral_High_Vol': {'min': -0.2, 'max': 0.3, 'vol_min': 0.6, 'confidence_min': 0.5},
                'Neutral_Med_Vol': {'min': -0.2, 'max': 0.3, 'vol_min': 0.3, 'vol_max': 0.6, 'confidence_min': 0.5},
                'Neutral_Low_Vol': {'min': -0.2, 'max': 0.3, 'vol_max': 0.3, 'confidence_min': 0.5},
                'Mild_Bearish_High_Vol': {'min': -0.7, 'max': -0.2, 'vol_min': 0.6, 'confidence_min': 0.6},
                'Mild_Bearish_Med_Vol': {'min': -0.7, 'max': -0.2, 'vol_min': 0.3, 'vol_max': 0.6, 'confidence_min': 0.6},
                'Mild_Bearish_Low_Vol': {'min': -0.7, 'max': -0.2, 'vol_max': 0.3, 'confidence_min': 0.6},
                'Strong_Bearish_High_Vol': {'min': -1.0, 'max': -0.7, 'vol_min': 0.6, 'confidence_min': 0.8},
                'Strong_Bearish_Med_Vol': {'min': -1.0, 'max': -0.7, 'vol_min': 0.3, 'vol_max': 0.6, 'confidence_min': 0.8},
                'Strong_Bearish_Low_Vol': {'min': -1.0, 'max': -0.7, 'vol_max': 0.3, 'confidence_min': 0.8},
                'Sideways_Consolidation': {'directional_max': 0.15, 'volatility_max': 0.25, 'confidence_min': 0.5},
                'Sideways_Breakout_Prep': {'directional_max': 0.25, 'volatility_min': 0.4, 'confidence_min': 0.6},
                'Transition_State': {'confidence_max': 0.4, 'stability_min': 0.3}
            }
        }
    
    def _get_dynamic_weight_configurations(self) -> Dict[str, Any]:
        """Get dynamic weighting parameters"""
        return {
            'Learning_Rate': {
                'value': 0.01,
                'description': 'Weight adjustment speed (0.001-0.1)',
                'range': '0.001-0.1'
            },
            'Adaptation_Period': {
                'value': 20,
                'description': 'Trading days for weight adaptation',
                'range': '10-50'
            },
            'Performance_Window': {
                'value': 100,
                'description': 'Data points for performance tracking',
                'range': '50-500'
            },
            'DTE_0_1_Weight_Reduction': {
                'value': 0.4,
                'description': '40% weight reduction for 0-1 DTE',
                'range': '0.2-0.6'
            },
            'DTE_2_3_Weight_Reduction': {
                'value': 0.2,
                'description': '20% weight reduction for 2-3 DTE',
                'range': '0.1-0.4'
            },
            'Historical_Decay_Factor': {
                'value': 0.95,
                'description': 'Performance decay factor',
                'range': '0.9-0.99'
            },
            'Min_Sample_Size': {
                'value': 30,
                'description': 'Minimum samples for weight adjustment',
                'range': '20-100'
            },
            'Confidence_Threshold': {
                'value': 0.6,
                'description': 'Minimum confidence for regime classification',
                'range': '0.4-0.8'
            },
            'Regime_Smoothing_Periods': {
                'value': 3,
                'description': 'Periods for regime smoothing',
                'range': '1-10'
            }
        }

    def create_indicator_configuration_sheet(self, workbook: openpyxl.Workbook, user_level: UserLevel = UserLevel.EXPERT) -> None:
        """Create enhanced IndicatorConfiguration sheet with detailed parameters"""
        ws = workbook.create_sheet("IndicatorConfiguration")

        # Enhanced headers with parameter columns
        headers = [
            'Indicator_Name', 'Enabled', 'Base_Weight', 'Min_Weight', 'Max_Weight',
            'Performance_Tracking', 'DTE_Adaptation', 'Timeframe_Support',
            'Parameter_1_Name', 'Parameter_1_Value', 'Parameter_1_Range', 'Parameter_1_Default',
            'Parameter_2_Name', 'Parameter_2_Value', 'Parameter_2_Range', 'Parameter_2_Default',
            'Parameter_3_Name', 'Parameter_3_Value', 'Parameter_3_Range', 'Parameter_3_Default',
            'Validation_Rules', 'Dependencies', 'Advanced_Mode_Only', 'Description'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add indicator data with enhanced parameters
        row = 2
        for indicator_name, config in self.indicator_configs.items():
            # Basic indicator configuration
            ws.cell(row=row, column=1, value=indicator_name)
            ws.cell(row=row, column=2, value=config['enabled'])
            ws.cell(row=row, column=3, value=config['base_weight'])
            ws.cell(row=row, column=4, value=config['min_weight'])
            ws.cell(row=row, column=5, value=config['max_weight'])
            ws.cell(row=row, column=6, value=config['performance_tracking'])
            ws.cell(row=row, column=7, value=config['dte_adaptation'])
            ws.cell(row=row, column=8, value=config['timeframe_support'])

            # Enhanced parameter configuration
            param_set = self.enhanced_params.get_indicator_parameters(indicator_name, user_level)
            if param_set:
                # Get top 3 most important parameters for this user level
                important_params = self._get_important_parameters(param_set, user_level, limit=3)

                for i, (param_name, param_def) in enumerate(important_params.items()):
                    if i < 3:  # Only show first 3 parameters
                        col_offset = 9 + (i * 4)  # Parameter columns start at column 9
                        ws.cell(row=row, column=col_offset, value=param_def.display_name)
                        ws.cell(row=row, column=col_offset + 1, value=param_def.default_value)

                        # Parameter range
                        if param_def.allowed_values:
                            range_str = f"Options: {', '.join(map(str, param_def.allowed_values))}"
                        elif param_def.min_value is not None and param_def.max_value is not None:
                            range_str = f"{param_def.min_value} - {param_def.max_value}"
                        else:
                            range_str = "Any"
                        ws.cell(row=row, column=col_offset + 2, value=range_str)
                        ws.cell(row=row, column=col_offset + 3, value=param_def.default_value)

                # Validation rules and dependencies
                validation_rules = []
                dependencies = []
                for param_def in param_set.parameters.values():
                    validation_rules.extend(param_def.validation_rules)
                    dependencies.extend(param_def.dependencies)

                ws.cell(row=row, column=21, value="; ".join(set(validation_rules)))
                ws.cell(row=row, column=22, value="; ".join(set(dependencies)))
                ws.cell(row=row, column=23, value=user_level == UserLevel.EXPERT)

            ws.cell(row=row, column=24, value=config['description'])
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _get_important_parameters(self, param_set, user_level: UserLevel, limit: int = 3) -> Dict[str, Any]:
        """Get most important parameters for user level"""
        # Sort parameters by impact level and user level appropriateness
        impact_order = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}

        filtered_params = {}
        for param_name, param_def in param_set.parameters.items():
            if user_level in param_def.user_levels:
                filtered_params[param_name] = param_def

        # Sort by impact level
        sorted_params = sorted(
            filtered_params.items(),
            key=lambda x: impact_order.get(x[1].impact_level, 0),
            reverse=True
        )

        # Return top parameters
        return dict(sorted_params[:limit])

    def create_detailed_indicator_parameters_sheet(self, workbook: openpyxl.Workbook, user_level: UserLevel = UserLevel.EXPERT) -> None:
        """Create detailed indicator parameters sheet"""
        ws = workbook.create_sheet("DetailedIndicatorParameters")

        # Headers
        headers = [
            'Indicator_Name', 'Parameter_Name', 'Display_Name', 'Parameter_Type',
            'Default_Value', 'Min_Value', 'Max_Value', 'Allowed_Values',
            'Current_Value', 'Description', 'Impact_Level', 'Category',
            'User_Level', 'Validation_Rules', 'Dependencies'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add parameter data
        row = 2
        for indicator_name in self.enhanced_params.get_all_indicators():
            param_set = self.enhanced_params.get_indicator_parameters(indicator_name, user_level)
            if not param_set:
                continue

            for param_name, param_def in param_set.parameters.items():
                ws.cell(row=row, column=1, value=indicator_name)
                ws.cell(row=row, column=2, value=param_name)
                ws.cell(row=row, column=3, value=param_def.display_name)
                ws.cell(row=row, column=4, value=param_def.parameter_type.value)
                ws.cell(row=row, column=5, value=str(param_def.default_value))
                ws.cell(row=row, column=6, value=param_def.min_value if param_def.min_value is not None else "")
                ws.cell(row=row, column=7, value=param_def.max_value if param_def.max_value is not None else "")
                ws.cell(row=row, column=8, value=", ".join(map(str, param_def.allowed_values)) if param_def.allowed_values else "")
                ws.cell(row=row, column=9, value=str(param_def.default_value))  # Current value starts as default
                ws.cell(row=row, column=10, value=param_def.description)
                ws.cell(row=row, column=11, value=param_def.impact_level)
                ws.cell(row=row, column=12, value=param_def.category)
                ws.cell(row=row, column=13, value=", ".join([ul.value for ul in param_def.user_levels]))
                ws.cell(row=row, column=14, value="; ".join(param_def.validation_rules))
                ws.cell(row=row, column=15, value="; ".join(param_def.dependencies))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_parameter_presets_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create parameter presets sheet"""
        ws = workbook.create_sheet("ParameterPresets")

        # Headers
        headers = [
            'Indicator_Name', 'Preset_Name', 'Parameter_Name', 'Parameter_Value',
            'Description', 'Use_Case', 'Risk_Level'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add preset data
        row = 2
        preset_descriptions = {
            'conservative': 'Lower risk, stable performance, suitable for risk-averse traders',
            'balanced': 'Moderate risk-reward balance, suitable for most trading strategies',
            'aggressive': 'Higher risk, potentially higher returns, suitable for experienced traders'
        }

        risk_levels = {
            'conservative': 'Low',
            'balanced': 'Medium',
            'aggressive': 'High'
        }

        for indicator_name in self.enhanced_params.get_all_indicators():
            param_set = self.enhanced_params.get_indicator_parameters(indicator_name)
            if not param_set:
                continue

            for preset_name, preset_values in param_set.presets.items():
                for param_name, param_value in preset_values.items():
                    ws.cell(row=row, column=1, value=indicator_name)
                    ws.cell(row=row, column=2, value=preset_name.title())
                    ws.cell(row=row, column=3, value=param_name)
                    ws.cell(row=row, column=4, value=str(param_value))
                    ws.cell(row=row, column=5, value=preset_descriptions.get(preset_name, ""))
                    ws.cell(row=row, column=6, value=f"{preset_name.title()} trading strategy")
                    ws.cell(row=row, column=7, value=risk_levels.get(preset_name, "Medium"))
                    row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_strike_configuration_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create StrikeConfiguration sheet"""
        ws = workbook.create_sheet("StrikeConfiguration")

        # Headers
        headers = ['Parameter', 'Value', 'Description', 'Options/Range']

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add strike configuration data
        row = 2
        for param_name, config in self.strike_configs.items():
            ws.cell(row=row, column=1, value=param_name)
            ws.cell(row=row, column=2, value=config['value'])
            ws.cell(row=row, column=3, value=config['description'])
            ws.cell(row=row, column=4, value=config.get('options', config.get('range', '')))
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_regime_formation_sheet(self, workbook: openpyxl.Workbook, regime_mode: str = "18_REGIME") -> None:
        """Create RegimeFormationRules sheet"""
        ws = workbook.create_sheet("RegimeFormationRules")

        # Add regime mode indicator
        ws.cell(row=1, column=1, value=f"Regime Mode: {regime_mode}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Headers based on regime mode
        if regime_mode == "8_REGIME":
            headers = ['Regime_Type', 'Score_Min', 'Score_Max', 'ATR_Min', 'ATR_Max',
                      'Directional_Max', 'Volatility_Max', 'Confidence_Min']
        else:  # 18_REGIME
            headers = ['Regime_Type', 'Score_Min', 'Score_Max', 'Vol_Min', 'Vol_Max',
                      'Directional_Max', 'Volatility_Min', 'Volatility_Max', 'Confidence_Min', 'Stability_Min']

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add regime data
        regime_data = self.regime_configs[regime_mode]
        row = 4
        for regime_name, config in regime_data.items():
            ws.cell(row=row, column=1, value=regime_name)
            ws.cell(row=row, column=2, value=config.get('min', ''))
            ws.cell(row=row, column=3, value=config.get('max', ''))

            if regime_mode == "8_REGIME":
                ws.cell(row=row, column=4, value=config.get('atr_min', ''))
                ws.cell(row=row, column=5, value=config.get('atr_max', ''))
                ws.cell(row=row, column=6, value=config.get('directional_max', ''))
                ws.cell(row=row, column=7, value=config.get('volatility_max', ''))
                ws.cell(row=row, column=8, value=config.get('confidence_min', ''))
            else:  # 18_REGIME
                ws.cell(row=row, column=4, value=config.get('vol_min', ''))
                ws.cell(row=row, column=5, value=config.get('vol_max', ''))
                ws.cell(row=row, column=6, value=config.get('directional_max', ''))
                ws.cell(row=row, column=7, value=config.get('volatility_min', ''))
                ws.cell(row=row, column=8, value=config.get('volatility_max', ''))
                ws.cell(row=row, column=9, value=config.get('confidence_min', ''))
                ws.cell(row=row, column=10, value=config.get('stability_min', ''))

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_dynamic_weight_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create DynamicWeightageParameters sheet"""
        ws = workbook.create_sheet("DynamicWeightageParameters")

        # Headers
        headers = ['Parameter', 'Value', 'Description', 'Range']

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add dynamic weight data
        row = 2
        for param_name, config in self.dynamic_weight_configs.items():
            ws.cell(row=row, column=1, value=param_name)
            ws.cell(row=row, column=2, value=config['value'])
            ws.cell(row=row, column=3, value=config['description'])
            ws.cell(row=row, column=4, value=config['range'])
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_timeframe_configuration_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create MultiTimeframeConfiguration sheet"""
        ws = workbook.create_sheet("MultiTimeframeConfiguration")

        # Headers
        headers = ['Timeframe', 'Enabled', 'Weight', 'Min_Confidence', 'Analysis_Type', 'Description']

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Timeframe configurations
        timeframes = [
            {'timeframe': '3min', 'enabled': True, 'weight': 0.15, 'min_confidence': 0.5,
             'analysis_type': 'SCALPING', 'description': 'High-frequency scalping analysis'},
            {'timeframe': '5min', 'enabled': True, 'weight': 0.25, 'min_confidence': 0.6,
             'analysis_type': 'INTRADAY', 'description': 'Primary intraday analysis'},
            {'timeframe': '10min', 'enabled': True, 'weight': 0.35, 'min_confidence': 0.7,
             'analysis_type': 'SWING', 'description': 'Short-term swing analysis'},
            {'timeframe': '15min', 'enabled': True, 'weight': 0.25, 'min_confidence': 0.8,
             'analysis_type': 'POSITION', 'description': 'Position-based analysis'},
            {'timeframe': '30min', 'enabled': False, 'weight': 0.0, 'min_confidence': 0.8,
             'analysis_type': 'TREND', 'description': 'Long-term trend analysis (optional)'}
        ]

        # Add timeframe data
        row = 2
        for tf_config in timeframes:
            ws.cell(row=row, column=1, value=tf_config['timeframe'])
            ws.cell(row=row, column=2, value=tf_config['enabled'])
            ws.cell(row=row, column=3, value=tf_config['weight'])
            ws.cell(row=row, column=4, value=tf_config['min_confidence'])
            ws.cell(row=row, column=5, value=tf_config['analysis_type'])
            ws.cell(row=row, column=6, value=tf_config['description'])
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_metadata_sheet(self, workbook: openpyxl.Workbook, template_type: str) -> None:
        """Create template metadata sheet"""
        ws = workbook.create_sheet("TemplateMetadata")

        # Template information
        metadata = [
            ['Template Type', template_type],
            ['Created Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Version', '1.0.0'],
            ['Compatible With', 'Backtester V2 Market Regime System'],
            ['Regime Modes', '8_REGIME, 18_REGIME'],
            ['Indicators Supported', '13 comprehensive indicators'],
            ['Strike Analysis', 'ATM, ITM1, OTM1, Combined'],
            ['Timeframes', '3min, 5min, 10min, 15min, 30min'],
            ['DTE Adaptation', 'Enabled with 0-3 and 4+ day configurations'],
            ['Dynamic Weights', 'Performance-based weight adjustment'],
            ['GPU Acceleration', 'Supported'],
            ['Real-time Streaming', 'WebSocket compatible'],
            ['Configuration Sheets', '6 comprehensive configuration sheets'],
            ['Usage Instructions', 'See documentation for detailed setup guide']
        ]

        # Add metadata
        for row, (key, value) in enumerate(metadata, 1):
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def generate_template(self, template_type: str = "18_REGIME", filename: Optional[str] = None, user_level: UserLevel = UserLevel.EXPERT) -> str:
        """
        Generate comprehensive market regime Excel template with enhanced parameters

        Args:
            template_type (str): Template type (8_REGIME, 18_REGIME, DEMO)
            filename (str, optional): Custom filename
            user_level (UserLevel): User experience level for parameter exposure

        Returns:
            str: Path to generated template file
        """
        try:
            # Create workbook
            workbook = openpyxl.Workbook()

            # Remove default sheet
            workbook.remove(workbook.active)

            # Create all configuration sheets with enhanced parameters
            self.create_indicator_configuration_sheet(workbook, user_level)
            self.create_detailed_indicator_parameters_sheet(workbook, user_level)
            self.create_parameter_presets_sheet(workbook)
            self.create_strike_configuration_sheet(workbook)
            self.create_regime_formation_sheet(workbook, template_type)
            self.create_dynamic_weight_sheet(workbook)
            self.create_timeframe_configuration_sheet(workbook)
            self.create_metadata_sheet(workbook, template_type)

            # Set active sheet to first configuration sheet
            workbook.active = workbook['IndicatorConfiguration']

            # Generate filename
            if not filename:
                filename = f"market_regime_{template_type.lower()}_config.xlsx"

            # Save template
            template_path = self.output_dir / filename
            workbook.save(template_path)

            logger.info(f"✅ Generated enhanced market regime template: {template_path}")
            logger.info(f"📊 Template includes {len(workbook.sheetnames)} sheets with detailed parameters")
            return str(template_path)

        except Exception as e:
            logger.error(f"❌ Failed to generate template: {e}")
            raise

    def generate_all_templates(self) -> Dict[str, str]:
        """Generate all template types with different user levels"""
        templates = {}

        try:
            # Generate expert-level templates (full parameter exposure)
            templates['8_REGIME_EXPERT'] = self.generate_template("8_REGIME", "market_regime_8_config.xlsx", UserLevel.EXPERT)
            templates['18_REGIME_EXPERT'] = self.generate_template("18_REGIME", "market_regime_18_config.xlsx", UserLevel.EXPERT)

            # Generate intermediate-level templates (moderate parameter exposure)
            templates['18_REGIME_INTERMEDIATE'] = self.generate_template("18_REGIME", "market_regime_intermediate_config.xlsx", UserLevel.INTERMEDIATE)

            # Generate novice-level templates (simplified parameter exposure)
            templates['18_REGIME_NOVICE'] = self.generate_template("18_REGIME", "market_regime_novice_config.xlsx", UserLevel.NOVICE)

            # Generate demo template (based on 18-regime intermediate level)
            templates['DEMO'] = self.generate_template("18_REGIME", "market_regime_demo_config.xlsx", UserLevel.INTERMEDIATE)

            # Generate default template (expert level)
            templates['DEFAULT'] = self.generate_template("18_REGIME", "market_regime_config.xlsx", UserLevel.EXPERT)

            logger.info(f"✅ Generated {len(templates)} market regime templates with different user levels")
            return templates

        except Exception as e:
            logger.error(f"❌ Failed to generate templates: {e}")
            raise


def main():
    """Main function to generate market regime templates"""
    try:
        # Initialize generator
        generator = MarketRegimeTemplateGenerator()

        # Generate all templates
        templates = generator.generate_all_templates()

        # Print results
        print("\n🎯 Market Regime Templates Generated Successfully!")
        print("=" * 60)
        for template_type, path in templates.items():
            print(f"📄 {template_type}: {path}")

        print("\n✅ All templates ready for use!")
        print("📋 Upload these templates via the Market Regime strategy interface")

    except Exception as e:
        print(f"❌ Template generation failed: {e}")
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
