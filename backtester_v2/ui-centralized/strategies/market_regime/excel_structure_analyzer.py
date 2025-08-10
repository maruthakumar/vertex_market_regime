#!/usr/bin/env python3
"""
Excel Structure Analyzer for Market Regime Formation System

This script analyzes the current Excel configuration files to document their
structure, parameter counts, and identify gaps for the comprehensive testing framework.

Author: The Augster
Date: 2025-06-19
Version: 1.0.0
"""

import os
import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional
import logging
from pathlib import Path
import json

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelStructureAnalyzer:
    """Analyze Excel configuration files structure and parameters"""

    def __init__(self, base_path: str):
        """Initialize analyzer with base path"""
        self.base_path = Path(base_path)
        self.analysis_results = {}

    def analyze_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single Excel file structure"""
        try:
            logger.info(f"Analyzing Excel file: {file_path.name}")

            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            file_analysis = {
                'file_name': file_path.name,
                'file_path': str(file_path),
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'sheets_analysis': {},
                'total_parameters': 0,
                'parameter_categories': {}
            }

            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = self.analyze_sheet(sheet, sheet_name)
                file_analysis['sheets_analysis'][sheet_name] = sheet_analysis
                file_analysis['total_parameters'] += sheet_analysis['parameter_count']

                # Categorize parameters
                for category, count in sheet_analysis['parameter_categories'].items():
                    if category not in file_analysis['parameter_categories']:
                        file_analysis['parameter_categories'][category] = 0
                    file_analysis['parameter_categories'][category] += count

            workbook.close()
            return file_analysis

        except Exception as e:
            logger.error(f"Error analyzing {file_path}: {e}")
            return {
                'file_name': file_path.name,
                'error': str(e),
                'analysis_failed': True
            }

    def analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze a single sheet structure"""
        try:
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Extract data
            data = []
            headers = []

            # Try to identify headers (first row with data)
            for row in sheet.iter_rows(min_row=1, max_row=min(10, max_row), values_only=True):
                if any(cell is not None for cell in row):
                    headers = [str(cell) if cell is not None else '' for cell in row]
                    break

            # Extract all data
            for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
                if any(cell is not None for cell in row):
                    data.append([str(cell) if cell is not None else '' for cell in row])

            # Analyze parameter types
            parameter_categories = self.categorize_parameters(headers, data)

            sheet_analysis = {
                'sheet_name': sheet_name,
                'dimensions': {'rows': max_row, 'columns': max_col},
                'headers': headers,
                'data_rows': len(data),
                'parameter_count': len([h for h in headers if h and h != '']),
                'parameter_categories': parameter_categories,
                'sample_data': data[:5] if data else []  # First 5 rows as sample
            }

            return sheet_analysis

        except Exception as e:
            logger.error(f"Error analyzing sheet {sheet_name}: {e}")
            return {
                'sheet_name': sheet_name,
                'error': str(e),
                'analysis_failed': True
            }

    def categorize_parameters(self, headers: List[str], data: List[List[str]]) -> Dict[str, int]:
        """Categorize parameters based on headers and data"""
        categories = {
            'weights': 0,
            'thresholds': 0,
            'timeframes': 0,
            'indicators': 0,
            'configuration': 0,
            'validation': 0,
            'other': 0
        }

        weight_keywords = ['weight', 'allocation', 'ratio', 'proportion']
        threshold_keywords = ['threshold', 'limit', 'min', 'max', 'range']
        timeframe_keywords = ['timeframe', 'period', 'window', 'interval', 'time']
        indicator_keywords = ['indicator', 'signal', 'analysis', 'calculation']
        config_keywords = ['config', 'setting', 'parameter', 'option', 'enable']
        validation_keywords = ['validation', 'check', 'verify', 'test', 'rule']

        for header in headers:
            if not header:
                continue

            header_lower = header.lower()
            categorized = False

            for keyword in weight_keywords:
                if keyword in header_lower:
                    categories['weights'] += 1
                    categorized = True
                    break

            if not categorized:
                for keyword in threshold_keywords:
                    if keyword in header_lower:
                        categories['thresholds'] += 1
                        categorized = True
                        break

            if not categorized:
                for keyword in timeframe_keywords:
                    if keyword in header_lower:
                        categories['timeframes'] += 1
                        categorized = True
                        break

            if not categorized:
                for keyword in indicator_keywords:
                    if keyword in header_lower:
                        categories['indicators'] += 1
                        categorized = True
                        break

            if not categorized:
                for keyword in config_keywords:
                    if keyword in header_lower:
                        categories['configuration'] += 1
                        categorized = True
                        break

            if not categorized:
                for keyword in validation_keywords:
                    if keyword in header_lower:
                        categories['validation'] += 1
                        categorized = True
                        break

            if not categorized:
                categories['other'] += 1

        return categories

    def find_excel_files(self) -> List[Path]:
        """Find all Excel files in the directory"""
        excel_files = []

        # Look for .xlsx files
        for file_path in self.base_path.glob("*.xlsx"):
            if file_path.is_file():
                excel_files.append(file_path)

        logger.info(f"Found {len(excel_files)} Excel files")
        return excel_files

    def analyze_all_files(self) -> Dict[str, Any]:
        """Analyze all Excel files in the directory"""
        excel_files = self.find_excel_files()

        comprehensive_analysis = {
            'analysis_timestamp': pd.Timestamp.now().isoformat(),
            'base_path': str(self.base_path),
            'total_files': len(excel_files),
            'files_analyzed': [],
            'summary': {
                'total_sheets': 0,
                'total_parameters': 0,
                'parameter_categories_summary': {},
                'configuration_gaps': []
            }
        }

        # Analyze each file
        for file_path in excel_files:
            file_analysis = self.analyze_excel_file(file_path)
            comprehensive_analysis['files_analyzed'].append(file_analysis)

            if not file_analysis.get('analysis_failed', False):
                comprehensive_analysis['summary']['total_sheets'] += file_analysis['sheet_count']
                comprehensive_analysis['summary']['total_parameters'] += file_analysis['total_parameters']

                # Aggregate parameter categories
                for category, count in file_analysis['parameter_categories'].items():
                    if category not in comprehensive_analysis['summary']['parameter_categories_summary']:
                        comprehensive_analysis['summary']['parameter_categories_summary'][category] = 0
                    comprehensive_analysis['summary']['parameter_categories_summary'][category] += count

        # Identify configuration gaps
        comprehensive_analysis['summary']['configuration_gaps'] = self.identify_configuration_gaps(
            comprehensive_analysis['files_analyzed']
        )

        return comprehensive_analysis

    def identify_configuration_gaps(self, files_analyzed: List[Dict[str, Any]]) -> List[str]:
        """Identify gaps in current configuration for 100% rolling Triple Straddle"""
        gaps = []

        # Required sheets for comprehensive testing
        required_sheets = [
            'TripleStraddleAnalysis',
            'GreekSentimentAnalysis',
            'TrendingOIWithPA',
            'IVAnalysis',
            'ATRTechnical',
            'RegimeClassification'
        ]

        # Required parameter categories
        required_parameters = {
            'rolling_configuration': ['rolling_weight', 'static_weight', 'confidence_weighting'],
            'component_weights': ['triple_straddle_weight', 'greek_sentiment_weight', 'trending_oi_weight'],
            'timeframe_analysis': ['3min_weight', '5min_weight', '10min_weight', '15min_weight'],
            'validation_rules': ['weight_sum_validation', 'range_validation', 'dependency_validation']
        }

        # Check for missing sheets
        all_sheets = []
        for file_analysis in files_analyzed:
            if not file_analysis.get('analysis_failed', False):
                all_sheets.extend(file_analysis['sheet_names'])

        for required_sheet in required_sheets:
            if not any(required_sheet.lower() in sheet.lower() for sheet in all_sheets):
                gaps.append(f"Missing sheet: {required_sheet}")

        # Check for 100% rolling configuration parameters
        rolling_params_found = False
        for file_analysis in files_analyzed:
            if not file_analysis.get('analysis_failed', False):
                for sheet_name, sheet_analysis in file_analysis['sheets_analysis'].items():
                    if not sheet_analysis.get('analysis_failed', False):
                        headers = sheet_analysis.get('headers', [])
                        if any('rolling' in str(header).lower() for header in headers):
                            rolling_params_found = True
                            break

        if not rolling_params_found:
            gaps.append("Missing 100% rolling configuration parameters")

        # Check for unified configuration structure
        unified_config_found = False
        for file_analysis in files_analyzed:
            if 'unified' in file_analysis['file_name'].lower():
                unified_config_found = True
                break

        if not unified_config_found:
            gaps.append("Missing unified configuration template")

        return gaps

    def generate_report(self, output_path: Optional[str] = None) -> str:
        """Generate comprehensive analysis report"""
        analysis_results = self.analyze_all_files()

        if output_path is None:
            output_path = self.base_path / "config_analysis_report.json"

        # Save detailed analysis as JSON
        with open(output_path, 'w') as f:
            json.dump(analysis_results, f, indent=2)

        # Generate markdown report
        report_path = str(output_path).replace('.json', '.md')
        self.generate_markdown_report(analysis_results, report_path)

        logger.info(f"Analysis complete. Reports saved to {output_path} and {report_path}")
        return report_path

    def generate_markdown_report(self, analysis_results: Dict[str, Any], report_path: str):
        """Generate markdown report"""
        with open(report_path, 'w') as f:
            f.write("# Excel Configuration Analysis Report\n\n")
            f.write(f"**Analysis Date:** {analysis_results['analysis_timestamp']}\n")
            f.write(f"**Base Path:** {analysis_results['base_path']}\n\n")

            # Summary
            f.write("## Summary\n\n")
            summary = analysis_results['summary']
            f.write(f"- **Total Files Analyzed:** {analysis_results['total_files']}\n")
            f.write(f"- **Total Sheets:** {summary['total_sheets']}\n")
            f.write(f"- **Total Parameters:** {summary['total_parameters']}\n\n")

            # Parameter Categories
            f.write("### Parameter Categories\n\n")
            for category, count in summary['parameter_categories_summary'].items():
                f.write(f"- **{category.title()}:** {count}\n")
            f.write("\n")

            # Configuration Gaps
            f.write("### Configuration Gaps Identified\n\n")
            for gap in summary['configuration_gaps']:
                f.write(f"- {gap}\n")
            f.write("\n")

            # Detailed File Analysis
            f.write("## Detailed File Analysis\n\n")
            for file_analysis in analysis_results['files_analyzed']:
                if file_analysis.get('analysis_failed', False):
                    f.write(f"### {file_analysis['file_name']} (FAILED)\n")
                    f.write(f"**Error:** {file_analysis.get('error', 'Unknown error')}\n\n")
                    continue

                f.write(f"### {file_analysis['file_name']}\n\n")
                f.write(f"- **Sheets:** {file_analysis['sheet_count']}\n")
                f.write(f"- **Total Parameters:** {file_analysis['total_parameters']}\n")
                f.write(f"- **Sheet Names:** {', '.join(file_analysis['sheet_names'])}\n\n")

                # Sheet details
                for sheet_name, sheet_analysis in file_analysis['sheets_analysis'].items():
                    if sheet_analysis.get('analysis_failed', False):
                        continue
                    f.write(f"#### Sheet: {sheet_name}\n")
                    f.write(f"- **Parameters:** {sheet_analysis['parameter_count']}\n")
                    f.write(f"- **Data Rows:** {sheet_analysis['data_rows']}\n")
                    f.write(f"- **Dimensions:** {sheet_analysis['dimensions']['rows']}x{sheet_analysis['dimensions']['columns']}\n\n")

if __name__ == "__main__":
    # Initialize analyzer
    base_path = "/srv/samba/shared/bt/backtester_stable/BTRUN/backtester_v2/market_regime"
    analyzer = ExcelStructureAnalyzer(base_path)

    # Generate comprehensive analysis report
    report_path = analyzer.generate_report()
    print(f"Analysis complete. Report saved to: {report_path}")