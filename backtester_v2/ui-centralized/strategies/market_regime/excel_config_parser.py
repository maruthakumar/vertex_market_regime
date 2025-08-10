"""
Market Regime Excel Configuration Parser
=======================================

This module provides comprehensive Excel configuration parsing for the Market Regime Detection System,
converting Excel templates into structured configuration objects for the enhanced market regime engine.

Features:
- Comprehensive Excel parsing with validation
- Support for all configuration sheets
- Error handling and detailed validation reporting
- Integration with enhanced market regime package
- Dynamic weight configuration parsing
- Multi-timeframe configuration support
- Strike analysis parameter extraction
- Regime formation rule parsing

Author: Market Regime Integration Team
Date: 2025-06-15
"""

import pandas as pd
import numpy as np
import os
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
from datetime import datetime
import openpyxl
from dataclasses import dataclass, field
import json

# Import configuration manager
try:
    from .config_manager import get_config_manager
    config_manager = get_config_manager()
except ImportError:
    # Fallback for standalone testing
    import sys
    from pathlib import Path
    sys.path.append(str(Path(__file__).parent))
    from config_manager import get_config_manager
    config_manager = get_config_manager()

# Import enhanced parameter system
try:
    from .archive_enhanced_modules_do_not_use.enhanced_indicator_parameters import EnhancedIndicatorParameters, UserLevel
except ImportError:
    # Handle relative import when running as script
    import sys
    from pathlib import Path
    current_dir = Path(__file__).parent
    sys.path.insert(0, str(current_dir))
    from enhanced_indicator_parameters import EnhancedIndicatorParameters, UserLevel

logger = logging.getLogger(__name__)

@dataclass
class IndicatorConfig:
    """Configuration for individual indicators"""
    name: str
    enabled: bool
    base_weight: float
    min_weight: float
    max_weight: float
    performance_tracking: bool
    dte_adaptation: bool
    timeframe_support: List[str]
    description: str

@dataclass
class StrikeConfig:
    """Strike analysis configuration"""
    atm_method: str = "CLOSEST_TO_UNDERLYING"
    otm1_offset_dte_0_3: int = 50
    otm1_offset_dte_4_plus: int = 100
    itm1_offset_dte_0_3: int = -50
    itm1_offset_dte_4_plus: int = -100
    strike_weighting_method: str = "EQUAL_WEIGHTED"
    combined_analysis_enabled: bool = True
    multi_timeframe_enabled: bool = True

@dataclass
class RegimeThreshold:
    """Individual regime threshold configuration"""
    min_score: Optional[float] = None
    max_score: Optional[float] = None
    vol_min: Optional[float] = None
    vol_max: Optional[float] = None
    directional_max: Optional[float] = None
    volatility_min: Optional[float] = None
    volatility_max: Optional[float] = None
    confidence_min: float = 0.6
    stability_min: Optional[float] = None

@dataclass
class DynamicWeightConfig:
    """Dynamic weighting configuration"""
    learning_rate: float = 0.01
    adaptation_period: int = 20
    performance_window: int = 100
    dte_0_1_weight_reduction: float = 0.4
    dte_2_3_weight_reduction: float = 0.2
    historical_decay_factor: float = 0.95
    min_sample_size: int = 30
    confidence_threshold: float = 0.6
    regime_smoothing_periods: int = 3

@dataclass
class TimeframeConfig:
    """Multi-timeframe configuration"""
    timeframe: str
    enabled: bool
    weight: float
    min_confidence: float
    analysis_type: str
    description: str

@dataclass
class MarketRegimeConfig:
    """Complete market regime configuration with enhanced parameters"""
    regime_mode: str = "18_REGIME"
    indicators: Dict[str, IndicatorConfig] = field(default_factory=dict)
    strike_config: StrikeConfig = field(default_factory=StrikeConfig)
    regime_thresholds: Dict[str, RegimeThreshold] = field(default_factory=dict)
    dynamic_weights: DynamicWeightConfig = field(default_factory=DynamicWeightConfig)
    timeframes: Dict[str, TimeframeConfig] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)
    # Enhanced parameter configurations
    detailed_parameters: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    parameter_presets: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    user_level: str = "expert"

class MarketRegimeExcelParser:
    """Parse Excel configuration files for Market Regime system"""
    
    def __init__(self):
        """Initialize Excel parser"""
        self.required_sheets = [
            'IndicatorConfiguration',
            'StrikeConfiguration',
            'RegimeFormationRules',
            'DynamicWeightageParameters',
            'MultiTimeframeConfiguration'
        ]
        self.optional_sheets = [
            'TemplateMetadata',
            'DetailedIndicatorParameters',
            'ParameterPresets'
        ]
        
        logger.info("✅ MarketRegimeExcelParser initialized")
    
    def validate_excel_file(self, excel_path: str) -> Tuple[bool, str, Optional[str]]:
        """
        Validate Excel file structure and content
        
        Args:
            excel_path (str): Path to Excel file
            
        Returns:
            Tuple[bool, str, Optional[str]]: (is_valid, error_message, regime_mode)
        """
        try:
            excel_path = Path(excel_path)
            if not excel_path.exists():
                return False, f"Excel file not found: {excel_path}", None
            
            # Load workbook
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet_names = workbook.sheetnames
            
            # Check required sheets
            missing_sheets = []
            for required_sheet in self.required_sheets:
                if required_sheet not in sheet_names:
                    missing_sheets.append(required_sheet)
            
            if missing_sheets:
                return False, f"Missing required sheets: {', '.join(missing_sheets)}", None
            
            # Detect regime mode from RegimeFormationRules sheet
            regime_mode = self._detect_regime_mode(workbook)
            
            # Validate sheet contents
            validation_errors = []
            
            # Validate IndicatorConfiguration
            indicator_errors = self._validate_indicator_sheet(workbook)
            if indicator_errors:
                validation_errors.extend(indicator_errors)
            
            # Validate StrikeConfiguration
            strike_errors = self._validate_strike_sheet(workbook)
            if strike_errors:
                validation_errors.extend(strike_errors)
            
            # Validate RegimeFormationRules
            regime_errors = self._validate_regime_sheet(workbook, regime_mode)
            if regime_errors:
                validation_errors.extend(regime_errors)
            
            if validation_errors:
                return False, f"Validation errors: {'; '.join(validation_errors)}", regime_mode
            
            logger.info(f"✅ Excel file validation successful: {excel_path}")
            return True, "Validation successful", regime_mode
            
        except Exception as e:
            logger.error(f"❌ Excel validation failed: {e}")
            return False, f"Validation failed: {str(e)}", None
    
    def _detect_regime_mode(self, workbook: openpyxl.Workbook) -> str:
        """Detect regime mode from RegimeFormationRules sheet"""
        try:
            ws = workbook['RegimeFormationRules']

            # Check first cell for regime mode indicator
            first_cell = ws.cell(row=1, column=1).value
            if first_cell and "Regime Mode:" in str(first_cell):
                if "8_REGIME" in str(first_cell):
                    return "8_REGIME"
                elif "12_REGIME" in str(first_cell):
                    return "12_REGIME"
                elif "18_REGIME" in str(first_cell):
                    return "18_REGIME"

            # Check for REGIME_COMPLEXITY setting
            for row in range(1, min(10, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "REGIME_COMPLEXITY" in str(cell_value):
                    complexity_value = ws.cell(row=row, column=2).value
                    if complexity_value:
                        if "8_REGIME" in str(complexity_value):
                            return "8_REGIME"
                        elif "12_REGIME" in str(complexity_value):
                            return "12_REGIME"
                        elif "18_REGIME" in str(complexity_value):
                            return "18_REGIME"

            # Count regime types to determine mode
            regime_count = 0
            for row in range(4, ws.max_row + 1):
                regime_name = ws.cell(row=row, column=1).value
                if regime_name and str(regime_name).strip():
                    regime_count += 1

            # Determine mode based on count
            if regime_count <= 10:
                return "8_REGIME"
            elif regime_count <= 14:
                return "12_REGIME"
            else:
                return "18_REGIME"

        except Exception as e:
            logger.warning(f"Could not detect regime mode, defaulting to 18_REGIME: {e}")
            return "18_REGIME"
    
    def _validate_indicator_sheet(self, workbook: openpyxl.Workbook) -> List[str]:
        """Validate IndicatorConfiguration sheet"""
        errors = []
        try:
            ws = workbook['IndicatorConfiguration']
            
            # Check minimum required headers (flexible for enhanced templates)
            required_headers = [
                'Indicator_Name', 'Enabled', 'Base_Weight', 'Min_Weight', 'Max_Weight',
                'Performance_Tracking', 'DTE_Adaptation', 'Timeframe_Support'
            ]

            # Get all actual headers
            actual_headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    actual_headers.append(str(header))

            # Check that all required headers are present (order doesn't matter)
            missing_headers = []
            for required_header in required_headers:
                if required_header not in actual_headers:
                    missing_headers.append(required_header)

            if missing_headers:
                errors.append(f"IndicatorConfiguration: Missing required headers: {', '.join(missing_headers)}")

            # Check for data rows
            if ws.max_row < 2:
                errors.append("IndicatorConfiguration: No data rows found")
            
            # Validate data rows
            for row in range(2, ws.max_row + 1):
                indicator_name = ws.cell(row=row, column=1).value
                if not indicator_name:
                    continue
                
                # Validate weights
                base_weight = ws.cell(row=row, column=3).value
                min_weight = ws.cell(row=row, column=4).value
                max_weight = ws.cell(row=row, column=5).value
                
                if not isinstance(base_weight, (int, float)) or base_weight < 0 or base_weight > 1:
                    errors.append(f"IndicatorConfiguration row {row}: Invalid base_weight for {indicator_name}")
                
                if not isinstance(min_weight, (int, float)) or min_weight < 0 or min_weight > 1:
                    errors.append(f"IndicatorConfiguration row {row}: Invalid min_weight for {indicator_name}")
                
                if not isinstance(max_weight, (int, float)) or max_weight < 0 or max_weight > 1:
                    errors.append(f"IndicatorConfiguration row {row}: Invalid max_weight for {indicator_name}")
                
                if isinstance(min_weight, (int, float)) and isinstance(max_weight, (int, float)):
                    if min_weight > max_weight:
                        errors.append(f"IndicatorConfiguration row {row}: min_weight > max_weight for {indicator_name}")
                
        except Exception as e:
            errors.append(f"IndicatorConfiguration validation error: {str(e)}")
        
        return errors
    
    def _validate_strike_sheet(self, workbook: openpyxl.Workbook) -> List[str]:
        """Validate StrikeConfiguration sheet"""
        errors = []
        try:
            ws = workbook['StrikeConfiguration']
            
            # Required parameters
            required_params = [
                'ATM_Method', 'OTM1_Offset_DTE_0_3', 'OTM1_Offset_DTE_4_Plus',
                'ITM1_Offset_DTE_0_3', 'ITM1_Offset_DTE_4_Plus', 'Strike_Weighting_Method'
            ]
            
            found_params = set()
            for row in range(2, ws.max_row + 1):
                param_name = ws.cell(row=row, column=1).value
                if param_name:
                    found_params.add(str(param_name))
            
            for required_param in required_params:
                if required_param not in found_params:
                    errors.append(f"StrikeConfiguration: Missing required parameter '{required_param}'")
                    
        except Exception as e:
            errors.append(f"StrikeConfiguration validation error: {str(e)}")
        
        return errors
    
    def _validate_regime_sheet(self, workbook: openpyxl.Workbook, regime_mode: str) -> List[str]:
        """Validate RegimeFormationRules sheet"""
        errors = []
        try:
            ws = workbook['RegimeFormationRules']
            
            # Count regime definitions
            regime_count = 0
            for row in range(4, ws.max_row + 1):
                regime_name = ws.cell(row=row, column=1).value
                if regime_name and str(regime_name).strip():
                    regime_count += 1
            
            # Validate regime count based on mode
            if regime_mode == "8_REGIME" and regime_count < 6:
                errors.append(f"RegimeFormationRules: Expected at least 6 regimes for 8_REGIME mode, found {regime_count}")
            elif regime_mode == "18_REGIME" and regime_count < 15:
                errors.append(f"RegimeFormationRules: Expected at least 15 regimes for 18_REGIME mode, found {regime_count}")
                
        except Exception as e:
            errors.append(f"RegimeFormationRules validation error: {str(e)}")

        return errors

    def parse_excel_config(self, excel_path: str) -> MarketRegimeConfig:
        """
        Parse Excel configuration file into structured configuration object

        Args:
            excel_path (str): Path to Excel configuration file

        Returns:
            MarketRegimeConfig: Parsed configuration object
        """
        try:
            # Validate file first
            is_valid, error_msg, regime_mode = self.validate_excel_file(excel_path)
            if not is_valid:
                raise ValueError(f"Excel validation failed: {error_msg}")

            # Load workbook
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # Parse all configuration sheets
            config = MarketRegimeConfig(regime_mode=regime_mode)

            # Parse indicators
            config.indicators = self._parse_indicator_config(workbook)

            # Parse strike configuration
            config.strike_config = self._parse_strike_config(workbook)

            # Parse regime thresholds
            config.regime_thresholds = self._parse_regime_thresholds(workbook, regime_mode)

            # Parse dynamic weights
            config.dynamic_weights = self._parse_dynamic_weights(workbook)

            # Parse timeframes
            config.timeframes = self._parse_timeframes(workbook)

            # Parse metadata if available
            config.metadata = self._parse_metadata(workbook)

            # Parse enhanced parameters if available
            config.detailed_parameters = self._parse_detailed_parameters(workbook)
            config.parameter_presets = self._parse_parameter_presets(workbook)

            logger.info(f"✅ Successfully parsed Excel configuration: {excel_path}")
            logger.info(f"📊 Parsed {len(config.indicators)} indicators with {len(config.detailed_parameters)} detailed parameter sets")
            return config

        except Exception as e:
            logger.error(f"❌ Failed to parse Excel configuration: {e}")
            raise

    def _parse_indicator_config(self, workbook: openpyxl.Workbook) -> Dict[str, IndicatorConfig]:
        """Parse IndicatorConfiguration sheet (flexible for enhanced templates)"""
        indicators = {}
        try:
            ws = workbook['IndicatorConfiguration']

            # Get header mapping for flexible parsing
            header_map = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[str(header)] = col

            for row in range(2, ws.max_row + 1):
                indicator_name = ws.cell(row=row, column=1).value
                if not indicator_name:
                    continue

                # Parse timeframe support (flexible column lookup)
                timeframe_col = header_map.get('Timeframe_Support', 8)
                timeframe_str = ws.cell(row=row, column=timeframe_col).value or ""
                timeframes = [tf.strip() for tf in str(timeframe_str).split(',') if tf.strip()]

                # Parse description (flexible column lookup)
                desc_col = header_map.get('Description', None)
                if desc_col:
                    description = str(ws.cell(row=row, column=desc_col).value or "")
                else:
                    # Try to find description in later columns for enhanced templates
                    description = ""
                    for col in range(20, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if header and 'Description' in str(header):
                            description = str(ws.cell(row=row, column=col).value or "")
                            break

                indicator_config = IndicatorConfig(
                    name=str(indicator_name),
                    enabled=bool(ws.cell(row=row, column=header_map.get('Enabled', 2)).value),
                    base_weight=float(ws.cell(row=row, column=header_map.get('Base_Weight', 3)).value or 0.0),
                    min_weight=float(ws.cell(row=row, column=header_map.get('Min_Weight', 4)).value or 0.0),
                    max_weight=float(ws.cell(row=row, column=header_map.get('Max_Weight', 5)).value or 1.0),
                    performance_tracking=bool(ws.cell(row=row, column=header_map.get('Performance_Tracking', 6)).value),
                    dte_adaptation=bool(ws.cell(row=row, column=header_map.get('DTE_Adaptation', 7)).value),
                    timeframe_support=timeframes,
                    description=description
                )

                indicators[indicator_name] = indicator_config

        except Exception as e:
            logger.error(f"Error parsing indicator configuration: {e}")
            raise

        return indicators

    def _parse_strike_config(self, workbook: openpyxl.Workbook) -> StrikeConfig:
        """Parse StrikeConfiguration sheet"""
        try:
            ws = workbook['StrikeConfiguration']

            # Create default config
            strike_config = StrikeConfig()

            # Parse parameters
            for row in range(2, ws.max_row + 1):
                param_name = ws.cell(row=row, column=1).value
                param_value = ws.cell(row=row, column=2).value

                if not param_name or param_value is None:
                    continue

                param_name = str(param_name)

                # Map parameters to config attributes
                if param_name == 'ATM_Method':
                    strike_config.atm_method = str(param_value)
                elif param_name == 'OTM1_Offset_DTE_0_3':
                    strike_config.otm1_offset_dte_0_3 = int(param_value)
                elif param_name == 'OTM1_Offset_DTE_4_Plus':
                    strike_config.otm1_offset_dte_4_plus = int(param_value)
                elif param_name == 'ITM1_Offset_DTE_0_3':
                    strike_config.itm1_offset_dte_0_3 = int(param_value)
                elif param_name == 'ITM1_Offset_DTE_4_Plus':
                    strike_config.itm1_offset_dte_4_plus = int(param_value)
                elif param_name == 'Strike_Weighting_Method':
                    strike_config.strike_weighting_method = str(param_value)
                elif param_name == 'Combined_Analysis_Enabled':
                    strike_config.combined_analysis_enabled = bool(param_value)
                elif param_name == 'Multi_Timeframe_Enabled':
                    strike_config.multi_timeframe_enabled = bool(param_value)

            return strike_config

        except Exception as e:
            logger.error(f"Error parsing strike configuration: {e}")
            raise

    def _parse_regime_thresholds(self, workbook: openpyxl.Workbook, regime_mode: str) -> Dict[str, RegimeThreshold]:
        """Parse RegimeFormationRules sheet"""
        thresholds = {}
        try:
            ws = workbook['RegimeFormationRules']

            for row in range(4, ws.max_row + 1):
                regime_name = ws.cell(row=row, column=1).value
                if not regime_name:
                    continue

                regime_name = str(regime_name)

                # Parse threshold values based on regime mode
                threshold = RegimeThreshold()

                # Common fields
                threshold.min_score = self._safe_float(ws.cell(row=row, column=2).value)
                threshold.max_score = self._safe_float(ws.cell(row=row, column=3).value)

                if regime_mode == "8_REGIME":
                    threshold.volatility_min = self._safe_float(ws.cell(row=row, column=4).value)  # ATR_Min
                    threshold.volatility_max = self._safe_float(ws.cell(row=row, column=5).value)  # ATR_Max
                    threshold.directional_max = self._safe_float(ws.cell(row=row, column=6).value)
                    threshold.confidence_min = self._safe_float(ws.cell(row=row, column=8).value) or 0.6
                else:  # 18_REGIME
                    threshold.vol_min = self._safe_float(ws.cell(row=row, column=4).value)
                    threshold.vol_max = self._safe_float(ws.cell(row=row, column=5).value)
                    threshold.directional_max = self._safe_float(ws.cell(row=row, column=6).value)
                    threshold.volatility_min = self._safe_float(ws.cell(row=row, column=7).value)
                    threshold.volatility_max = self._safe_float(ws.cell(row=row, column=8).value)
                    threshold.confidence_min = self._safe_float(ws.cell(row=row, column=9).value) or 0.6
                    threshold.stability_min = self._safe_float(ws.cell(row=row, column=10).value)

                thresholds[regime_name] = threshold

        except Exception as e:
            logger.error(f"Error parsing regime thresholds: {e}")
            raise

        return thresholds

    def _parse_dynamic_weights(self, workbook: openpyxl.Workbook) -> DynamicWeightConfig:
        """Parse DynamicWeightageParameters sheet"""
        try:
            ws = workbook['DynamicWeightageParameters']

            # Create default config
            dynamic_config = DynamicWeightConfig()

            # Parse parameters
            for row in range(2, ws.max_row + 1):
                param_name = ws.cell(row=row, column=1).value
                param_value = ws.cell(row=row, column=2).value

                if not param_name or param_value is None:
                    continue

                param_name = str(param_name)

                # Map parameters to config attributes
                if param_name == 'Learning_Rate':
                    dynamic_config.learning_rate = float(param_value)
                elif param_name == 'Adaptation_Period':
                    dynamic_config.adaptation_period = int(param_value)
                elif param_name == 'Performance_Window':
                    dynamic_config.performance_window = int(param_value)
                elif param_name == 'DTE_0_1_Weight_Reduction':
                    dynamic_config.dte_0_1_weight_reduction = float(param_value)
                elif param_name == 'DTE_2_3_Weight_Reduction':
                    dynamic_config.dte_2_3_weight_reduction = float(param_value)
                elif param_name == 'Historical_Decay_Factor':
                    dynamic_config.historical_decay_factor = float(param_value)
                elif param_name == 'Min_Sample_Size':
                    dynamic_config.min_sample_size = int(param_value)
                elif param_name == 'Confidence_Threshold':
                    dynamic_config.confidence_threshold = float(param_value)
                elif param_name == 'Regime_Smoothing_Periods':
                    dynamic_config.regime_smoothing_periods = int(param_value)

            return dynamic_config

        except Exception as e:
            logger.error(f"Error parsing dynamic weights: {e}")
            raise

    def _parse_timeframes(self, workbook: openpyxl.Workbook) -> Dict[str, TimeframeConfig]:
        """Parse MultiTimeframeConfiguration sheet"""
        timeframes = {}
        try:
            if 'MultiTimeframeConfiguration' not in workbook.sheetnames:
                # Return default timeframes if sheet doesn't exist
                return self._get_default_timeframes()

            ws = workbook['MultiTimeframeConfiguration']

            for row in range(2, ws.max_row + 1):
                timeframe = ws.cell(row=row, column=1).value
                if not timeframe:
                    continue

                timeframe_config = TimeframeConfig(
                    timeframe=str(timeframe),
                    enabled=bool(ws.cell(row=row, column=2).value),
                    weight=float(ws.cell(row=row, column=3).value or 0.0),
                    min_confidence=float(ws.cell(row=row, column=4).value or 0.5),
                    analysis_type=str(ws.cell(row=row, column=5).value or ""),
                    description=str(ws.cell(row=row, column=6).value or "")
                )

                timeframes[timeframe] = timeframe_config

        except Exception as e:
            logger.error(f"Error parsing timeframes: {e}")
            # Return default timeframes on error
            return self._get_default_timeframes()

        return timeframes

    def _parse_metadata(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Parse TemplateMetadata sheet if available"""
        metadata = {}
        try:
            if 'TemplateMetadata' not in workbook.sheetnames:
                return metadata

            ws = workbook['TemplateMetadata']

            for row in range(1, ws.max_row + 1):
                key = ws.cell(row=row, column=1).value
                value = ws.cell(row=row, column=2).value

                if key and value:
                    metadata[str(key)] = str(value)

        except Exception as e:
            logger.warning(f"Error parsing metadata: {e}")

        return metadata

    def _safe_float(self, value) -> Optional[float]:
        """Safely convert value to float"""
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def _get_default_timeframes(self) -> Dict[str, TimeframeConfig]:
        """Get default timeframe configurations"""
        return {
            '3min': TimeframeConfig('3min', True, 0.15, 0.5, 'SCALPING', 'High-frequency scalping analysis'),
            '5min': TimeframeConfig('5min', True, 0.25, 0.6, 'INTRADAY', 'Primary intraday analysis'),
            '10min': TimeframeConfig('10min', True, 0.35, 0.7, 'SWING', 'Short-term swing analysis'),
            '15min': TimeframeConfig('15min', True, 0.25, 0.8, 'POSITION', 'Position-based analysis')
        }

    def _parse_detailed_parameters(self, workbook: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
        """Parse DetailedIndicatorParameters sheet if available"""
        detailed_params = {}
        try:
            if 'DetailedIndicatorParameters' not in workbook.sheetnames:
                return detailed_params

            ws = workbook['DetailedIndicatorParameters']

            for row in range(2, ws.max_row + 1):
                indicator_name = ws.cell(row=row, column=1).value
                parameter_name = ws.cell(row=row, column=2).value

                if not indicator_name or not parameter_name:
                    continue

                if indicator_name not in detailed_params:
                    detailed_params[indicator_name] = {}

                # Parse parameter details
                param_details = {
                    'display_name': ws.cell(row=row, column=3).value or parameter_name,
                    'parameter_type': ws.cell(row=row, column=4).value or 'string',
                    'default_value': ws.cell(row=row, column=5).value,
                    'min_value': self._safe_float(ws.cell(row=row, column=6).value),
                    'max_value': self._safe_float(ws.cell(row=row, column=7).value),
                    'allowed_values': self._parse_allowed_values(ws.cell(row=row, column=8).value),
                    'current_value': ws.cell(row=row, column=9).value,
                    'description': ws.cell(row=row, column=10).value or '',
                    'impact_level': ws.cell(row=row, column=11).value or 'medium',
                    'category': ws.cell(row=row, column=12).value or 'general',
                    'user_levels': self._parse_user_levels(ws.cell(row=row, column=13).value),
                    'validation_rules': self._parse_list_field(ws.cell(row=row, column=14).value),
                    'dependencies': self._parse_list_field(ws.cell(row=row, column=15).value)
                }

                detailed_params[indicator_name][parameter_name] = param_details

        except Exception as e:
            logger.warning(f"Error parsing detailed parameters: {e}")

        return detailed_params

    def _parse_parameter_presets(self, workbook: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
        """Parse ParameterPresets sheet if available"""
        presets = {}
        try:
            if 'ParameterPresets' not in workbook.sheetnames:
                return presets

            ws = workbook['ParameterPresets']

            for row in range(2, ws.max_row + 1):
                indicator_name = ws.cell(row=row, column=1).value
                preset_name = ws.cell(row=row, column=2).value
                parameter_name = ws.cell(row=row, column=3).value
                parameter_value = ws.cell(row=row, column=4).value

                if not all([indicator_name, preset_name, parameter_name]):
                    continue

                # Create nested structure: presets[indicator][preset][parameter] = value
                if indicator_name not in presets:
                    presets[indicator_name] = {}

                if preset_name not in presets[indicator_name]:
                    presets[indicator_name][preset_name] = {
                        'parameters': {},
                        'description': ws.cell(row=row, column=5).value or '',
                        'use_case': ws.cell(row=row, column=6).value or '',
                        'risk_level': ws.cell(row=row, column=7).value or 'Medium'
                    }

                # Convert parameter value to appropriate type
                converted_value = self._convert_parameter_value(parameter_value)
                presets[indicator_name][preset_name]['parameters'][parameter_name] = converted_value

        except Exception as e:
            logger.warning(f"Error parsing parameter presets: {e}")

        return presets

    def _parse_allowed_values(self, value_str: str) -> Optional[List[str]]:
        """Parse allowed values from string"""
        if not value_str or value_str.strip() == "":
            return None

        # Split by comma and clean up
        values = [v.strip() for v in str(value_str).split(',') if v.strip()]
        return values if values else None

    def _parse_user_levels(self, levels_str: str) -> List[str]:
        """Parse user levels from string"""
        if not levels_str:
            return ['expert']

        levels = [level.strip().lower() for level in str(levels_str).split(',') if level.strip()]
        return levels if levels else ['expert']

    def _parse_list_field(self, field_str: str) -> List[str]:
        """Parse semicolon-separated list field"""
        if not field_str or field_str.strip() == "":
            return []

        items = [item.strip() for item in str(field_str).split(';') if item.strip()]
        return items

    def _convert_parameter_value(self, value: Any) -> Any:
        """Convert parameter value to appropriate Python type"""
        if value is None:
            return None

        value_str = str(value).strip().lower()

        # Boolean conversion
        if value_str in ['true', 'false']:
            return value_str == 'true'

        # Try numeric conversion
        try:
            if '.' in value_str:
                return float(value)
            else:
                return int(value)
        except (ValueError, TypeError):
            pass

        # Return as string
        return str(value)

    def export_config_to_json(self, config: MarketRegimeConfig, output_path: str) -> None:
        """Export configuration to JSON format for debugging"""
        try:
            # Convert config to dictionary
            config_dict = {
                'regime_mode': config.regime_mode,
                'indicators': {name: {
                    'name': ind.name,
                    'enabled': ind.enabled,
                    'base_weight': ind.base_weight,
                    'min_weight': ind.min_weight,
                    'max_weight': ind.max_weight,
                    'performance_tracking': ind.performance_tracking,
                    'dte_adaptation': ind.dte_adaptation,
                    'timeframe_support': ind.timeframe_support,
                    'description': ind.description
                } for name, ind in config.indicators.items()},
                'strike_config': {
                    'atm_method': config.strike_config.atm_method,
                    'otm1_offset_dte_0_3': config.strike_config.otm1_offset_dte_0_3,
                    'otm1_offset_dte_4_plus': config.strike_config.otm1_offset_dte_4_plus,
                    'itm1_offset_dte_0_3': config.strike_config.itm1_offset_dte_0_3,
                    'itm1_offset_dte_4_plus': config.strike_config.itm1_offset_dte_4_plus,
                    'strike_weighting_method': config.strike_config.strike_weighting_method,
                    'combined_analysis_enabled': config.strike_config.combined_analysis_enabled,
                    'multi_timeframe_enabled': config.strike_config.multi_timeframe_enabled
                },
                'metadata': config.metadata
            }

            # Save to JSON
            with open(output_path, 'w') as f:
                json.dump(config_dict, f, indent=2)

            logger.info(f"✅ Configuration exported to JSON: {output_path}")

        except Exception as e:
            logger.error(f"❌ Failed to export configuration to JSON: {e}")
            raise


def main():
    """Test function for Excel parser"""
    try:
        # Initialize parser
        parser = MarketRegimeExcelParser()

        # Test with sample file (if exists)
        test_file = os.path.join(config_manager.paths.get_input_sheets_path(), "market_regime_18_config.xlsx")

        if Path(test_file).exists():
            print(f"🧪 Testing Excel parser with: {test_file}")

            # Validate file
            is_valid, error_msg, regime_mode = parser.validate_excel_file(test_file)
            print(f"📋 Validation: {'✅ PASSED' if is_valid else '❌ FAILED'}")
            if not is_valid:
                print(f"   Error: {error_msg}")
            else:
                print(f"   Regime Mode: {regime_mode}")

                # Parse configuration
                config = parser.parse_excel_config(test_file)
                print(f"📊 Parsed {len(config.indicators)} indicators")
                print(f"🎯 Regime mode: {config.regime_mode}")
                print(f"⚙️ Strike config: {config.strike_config.atm_method}")

                # Export to JSON for inspection
                json_path = test_file.replace('.xlsx', '_parsed.json')
                parser.export_config_to_json(config, json_path)
                print(f"💾 Configuration exported to: {json_path}")
        else:
            print(f"⚠️ Test file not found: {test_file}")
            print("   Generate templates first using excel_template_generator.py")

    except Exception as e:
        print(f"❌ Parser test failed: {e}")
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
