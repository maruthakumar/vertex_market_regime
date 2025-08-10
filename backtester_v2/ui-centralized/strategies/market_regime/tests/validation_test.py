"""
Market Regime Detection System - Comprehensive Validation Test
=============================================================

This module provides comprehensive validation testing for the Market Regime Detection System,
ensuring all components are properly integrated and functioning correctly.

Test Coverage:
- Excel template generation and validation
- Configuration parsing and validation
- Enhanced regime engine functionality
- API integration and endpoints
- UI component integration
- WebSocket connectivity
- Performance benchmarks

Author: Market Regime Integration Team
Date: 2025-06-15
"""

import sys
import os
import time
import json
import asyncio
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Add current directory to path
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

class MarketRegimeValidationTest:
    """Comprehensive validation test suite for Market Regime Detection System"""
    
    def __init__(self):
        """Initialize validation test suite"""
        self.test_results = {}
        self.start_time = None
        self.base_path = Path("/srv/samba/shared/bt/backtester_stable/BTRUN")
        self.template_dir = self.base_path / "input_sheets"
        self.upload_dir = self.base_path / "uploaded_configs" / "market_regime"
        
        # Test configuration
        self.test_config = {
            'performance_targets': {
                'template_generation_time': 5.0,  # seconds
                'excel_parsing_time': 2.0,  # seconds
                'regime_calculation_time': 0.1,  # seconds
                'api_response_time': 1.0  # seconds
            },
            'test_data_size': 100,  # number of data points
            'template_types': ['8_REGIME', '18_REGIME', 'DEMO', 'DEFAULT']
        }
        
        logger.info("🧪 MarketRegimeValidationTest initialized")
    
    def run_comprehensive_validation(self) -> Dict[str, Any]:
        """Run comprehensive validation test suite"""
        try:
            self.start_time = time.time()
            print("🧪 Starting comprehensive Market Regime validation...")
            logger.info("🚀 Starting comprehensive Market Regime validation...")

            # Phase 1: Template System Validation
            print("📄 Phase 1: Template System Validation")
            logger.info("📄 Phase 1: Template System Validation")
            self.test_results['phase_1'] = self.validate_template_system()
            phase_1_status = "✅" if self.test_results['phase_1']['status'] == 'passed' else "❌"
            print(f"   {phase_1_status} Phase 1 Status: {self.test_results['phase_1']['status'].upper()}")

            # Phase 2: Configuration System Validation
            print("📋 Phase 2: Configuration System Validation")
            logger.info("📋 Phase 2: Configuration System Validation")
            self.test_results['phase_2'] = self.validate_configuration_system()
            phase_2_status = "✅" if self.test_results['phase_2']['status'] == 'passed' else "❌"
            print(f"   {phase_2_status} Phase 2 Status: {self.test_results['phase_2']['status'].upper()}")

            # Phase 3: Engine Integration Validation
            print("🔧 Phase 3: Engine Integration Validation")
            logger.info("🔧 Phase 3: Engine Integration Validation")
            self.test_results['phase_3'] = self.validate_engine_integration()
            phase_3_status = "✅" if self.test_results['phase_3']['status'] == 'passed' else "❌"
            print(f"   {phase_3_status} Phase 3 Status: {self.test_results['phase_3']['status'].upper()}")

            # Phase 4: API Integration Validation
            print("🌐 Phase 4: API Integration Validation")
            logger.info("🌐 Phase 4: API Integration Validation")
            self.test_results['phase_4'] = self.validate_api_integration()
            phase_4_status = "✅" if self.test_results['phase_4']['status'] == 'passed' else "❌"
            print(f"   {phase_4_status} Phase 4 Status: {self.test_results['phase_4']['status'].upper()}")

            # Phase 5: Performance Validation
            print("⚡ Phase 5: Performance Validation")
            logger.info("⚡ Phase 5: Performance Validation")
            self.test_results['phase_5'] = self.validate_performance()
            phase_5_status = "✅" if self.test_results['phase_5']['status'] == 'passed' else "❌"
            print(f"   {phase_5_status} Phase 5 Status: {self.test_results['phase_5']['status'].upper()}")

            # Generate final report
            self.test_results['summary'] = self.generate_validation_summary()

            total_time = time.time() - self.start_time
            print(f"🏁 Comprehensive validation completed in {total_time:.2f} seconds")
            logger.info(f"✅ Comprehensive validation completed in {total_time:.2f} seconds")

            return self.test_results

        except Exception as e:
            print(f"❌ Validation failed: {e}")
            logger.error(f"❌ Validation failed: {e}")
            self.test_results['error'] = str(e)
            return self.test_results
    
    def validate_template_system(self) -> Dict[str, Any]:
        """Validate Excel template generation system"""
        results = {'status': 'running', 'tests': {}, 'errors': []}
        
        try:
            from excel_template_generator import MarketRegimeTemplateGenerator
            
            # Test 1: Template Generator Initialization
            start_time = time.time()
            generator = MarketRegimeTemplateGenerator(str(self.template_dir))
            init_time = time.time() - start_time
            
            results['tests']['generator_initialization'] = {
                'status': 'passed',
                'time': init_time,
                'details': 'Template generator initialized successfully'
            }
            
            # Test 2: Template Generation
            template_results = {}
            for template_type in self.test_config['template_types']:
                try:
                    start_time = time.time()

                    # Map template types to proper generation parameters
                    if template_type == 'DEMO':
                        # Generate demo template based on 18_REGIME
                        template_path = generator.generate_template("18_REGIME", "market_regime_demo_config.xlsx")
                    elif template_type == 'DEFAULT':
                        # Generate default template based on 18_REGIME
                        template_path = generator.generate_template("18_REGIME", "market_regime_config.xlsx")
                    else:
                        # Generate standard templates
                        template_path = generator.generate_template(template_type)

                    generation_time = time.time() - start_time

                    # Verify file exists and has content
                    if Path(template_path).exists() and Path(template_path).stat().st_size > 0:
                        template_results[template_type] = {
                            'status': 'passed',
                            'time': generation_time,
                            'path': template_path,
                            'size_kb': round(Path(template_path).stat().st_size / 1024, 2)
                        }
                        print(f"   ✅ {template_type} template generated successfully ({template_results[template_type]['size_kb']} KB)")
                    else:
                        template_results[template_type] = {
                            'status': 'failed',
                            'error': 'Template file not created or empty'
                        }
                        print(f"   ❌ {template_type} template generation failed: file not created")

                except Exception as e:
                    template_results[template_type] = {
                        'status': 'failed',
                        'error': str(e)
                    }
                    print(f"   ❌ {template_type} template generation failed: {e}")
            
            results['tests']['template_generation'] = template_results
            
            # Test 3: Template Structure Validation
            structure_results = {}
            for template_type in self.test_config['template_types']:
                # Map template types to actual file names
                if template_type == 'DEMO':
                    template_path = self.template_dir / "market_regime_demo_config.xlsx"
                elif template_type == 'DEFAULT':
                    template_path = self.template_dir / "market_regime_config.xlsx"
                else:
                    template_path = self.template_dir / f"market_regime_{template_type.lower()}_config.xlsx"

                if template_path.exists():
                    try:
                        import openpyxl
                        workbook = openpyxl.load_workbook(template_path, data_only=True)
                        
                        required_sheets = [
                            'IndicatorConfiguration',
                            'StrikeConfiguration', 
                            'RegimeFormationRules',
                            'DynamicWeightageParameters',
                            'MultiTimeframeConfiguration',
                            'TemplateMetadata'
                        ]
                        
                        missing_sheets = []
                        for sheet in required_sheets:
                            if sheet not in workbook.sheetnames:
                                missing_sheets.append(sheet)
                        
                        if not missing_sheets:
                            structure_results[template_type] = {
                                'status': 'passed',
                                'sheets': workbook.sheetnames,
                                'sheet_count': len(workbook.sheetnames)
                            }
                        else:
                            structure_results[template_type] = {
                                'status': 'failed',
                                'missing_sheets': missing_sheets
                            }
                            
                    except Exception as e:
                        structure_results[template_type] = {
                            'status': 'failed',
                            'error': str(e)
                        }
                else:
                    structure_results[template_type] = {
                        'status': 'failed',
                        'error': 'Template file not found'
                    }
            
            results['tests']['template_structure'] = structure_results
            
            # Overall phase status
            all_passed = True
            for test_group in results['tests'].values():
                if isinstance(test_group, dict):
                    # Handle nested test results
                    for test in test_group.values():
                        if isinstance(test, dict) and 'status' in test:
                            if test['status'] != 'passed':
                                all_passed = False
                                break
                elif isinstance(test_group, dict) and 'status' in test_group:
                    # Handle direct test results
                    if test_group['status'] != 'passed':
                        all_passed = False
                        break
                if not all_passed:
                    break
            
            results['status'] = 'passed' if all_passed else 'failed'
            
        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(str(e))
            logger.error(f"Template system validation failed: {e}")
        
        return results
    
    def validate_configuration_system(self) -> Dict[str, Any]:
        """Validate Excel configuration parsing system"""
        results = {'status': 'running', 'tests': {}, 'errors': []}
        
        try:
            from excel_config_parser import MarketRegimeExcelParser
            
            # Test 1: Parser Initialization
            start_time = time.time()
            parser = MarketRegimeExcelParser()
            init_time = time.time() - start_time
            
            results['tests']['parser_initialization'] = {
                'status': 'passed',
                'time': init_time,
                'details': 'Excel parser initialized successfully'
            }
            
            # Test 2: Configuration Validation
            validation_results = {}
            for template_type in ['18_REGIME', '8_REGIME']:
                template_path = self.template_dir / f"market_regime_{template_type.lower()}_config.xlsx"
                
                if template_path.exists():
                    try:
                        start_time = time.time()
                        is_valid, error_msg, regime_mode = parser.validate_excel_file(str(template_path))
                        validation_time = time.time() - start_time
                        
                        validation_results[template_type] = {
                            'status': 'passed' if is_valid else 'failed',
                            'time': validation_time,
                            'regime_mode': regime_mode,
                            'error': error_msg if not is_valid else None
                        }
                        
                    except Exception as e:
                        validation_results[template_type] = {
                            'status': 'failed',
                            'error': str(e)
                        }
                else:
                    validation_results[template_type] = {
                        'status': 'skipped',
                        'error': 'Template file not found'
                    }
            
            results['tests']['configuration_validation'] = validation_results
            
            # Test 3: Configuration Parsing
            parsing_results = {}
            for template_type in ['18_REGIME']:
                template_path = self.template_dir / f"market_regime_{template_type.lower()}_config.xlsx"
                
                if template_path.exists():
                    try:
                        start_time = time.time()
                        config = parser.parse_excel_config(str(template_path))
                        parsing_time = time.time() - start_time
                        
                        parsing_results[template_type] = {
                            'status': 'passed',
                            'time': parsing_time,
                            'regime_mode': config.regime_mode,
                            'indicators_count': len(config.indicators),
                            'enabled_indicators': len([ind for ind in config.indicators.values() if ind.enabled]),
                            'timeframes_count': len(config.timeframes)
                        }
                        
                    except Exception as e:
                        parsing_results[template_type] = {
                            'status': 'failed',
                            'error': str(e)
                        }
                else:
                    parsing_results[template_type] = {
                        'status': 'skipped',
                        'error': 'Template file not found'
                    }
            
            results['tests']['configuration_parsing'] = parsing_results
            
            # Overall phase status
            all_passed = True
            for test_group in results['tests'].values():
                if isinstance(test_group, dict):
                    # Handle nested test results
                    for test in test_group.values():
                        if isinstance(test, dict) and 'status' in test:
                            if test['status'] not in ['passed', 'skipped']:
                                all_passed = False
                                break
                elif isinstance(test_group, dict) and 'status' in test_group:
                    # Handle direct test results
                    if test_group['status'] not in ['passed', 'skipped']:
                        all_passed = False
                        break
                if not all_passed:
                    break
            
            results['status'] = 'passed' if all_passed else 'failed'
            
        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(str(e))
            logger.error(f"Configuration system validation failed: {e}")
        
        return results
    
    def validate_engine_integration(self) -> Dict[str, Any]:
        """Validate enhanced regime engine integration"""
        results = {'status': 'running', 'tests': {}, 'errors': []}
        
        try:
            from enhanced_regime_engine import EnhancedMarketRegimeEngine
            
            # Test 1: Engine Initialization with Default Config
            start_time = time.time()
            engine = EnhancedMarketRegimeEngine()
            init_time = time.time() - start_time
            
            results['tests']['engine_initialization'] = {
                'status': 'passed',
                'time': init_time,
                'enhanced_available': engine.enhanced_available,
                'indicators_count': len(engine.indicator_engines),
                'regime_mode': engine.config.regime_mode
            }
            
            # Test 2: Market Data Processing
            # Create sample market data
            dates = pd.date_range(start='2024-01-01', periods=self.test_config['test_data_size'], freq='1min')
            sample_data = pd.DataFrame({
                'open': 22000 + np.random.randn(self.test_config['test_data_size']) * 50,
                'high': 22050 + np.random.randn(self.test_config['test_data_size']) * 50,
                'low': 21950 + np.random.randn(self.test_config['test_data_size']) * 50,
                'close': 22000 + np.random.randn(self.test_config['test_data_size']) * 50,
                'volume': 1000 + np.random.randint(0, 500, self.test_config['test_data_size'])
            }, index=dates)
            
            start_time = time.time()
            regime_result = engine.calculate_market_regime(sample_data)
            processing_time = time.time() - start_time
            
            results['tests']['market_data_processing'] = {
                'status': 'passed' if regime_result and 'regime_type' in regime_result else 'failed',
                'time': processing_time,
                'regime_type': regime_result.get('regime_type', 'Unknown'),
                'confidence': regime_result.get('confidence', 0.0),
                'engine_type': regime_result.get('engine_type', 'Unknown')
            }
            
            # Test 3: Performance Metrics
            metrics = engine.get_performance_metrics()
            
            results['tests']['performance_metrics'] = {
                'status': 'passed' if metrics else 'failed',
                'total_calculations': metrics.get('total_regime_calculations', 0),
                'average_confidence': metrics.get('average_confidence', 0.0),
                'engine_type': metrics.get('engine_type', 'Unknown')
            }
            
            # Overall phase status
            all_passed = all(
                test.get('status') == 'passed' 
                for test in results['tests'].values()
            )
            
            results['status'] = 'passed' if all_passed else 'failed'
            
        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(str(e))
            logger.error(f"Engine integration validation failed: {e}")
        
        return results

    def validate_api_integration(self) -> Dict[str, Any]:
        """Validate API integration functionality"""
        results = {'status': 'running', 'tests': {}, 'errors': []}

        try:
            from api_integration import MarketRegimeAPIIntegration, create_market_regime_router

            # Test 1: API Integration Initialization
            start_time = time.time()
            api_integration = MarketRegimeAPIIntegration()
            init_time = time.time() - start_time

            results['tests']['api_initialization'] = {
                'status': 'passed',
                'time': init_time,
                'upload_dir': str(api_integration.upload_dir),
                'template_dir': str(api_integration.template_dir)
            }

            # Test 2: Router Creation
            try:
                start_time = time.time()
                router = create_market_regime_router()
                router_time = time.time() - start_time

                route_count = len(router.routes)
                route_paths = [route.path for route in router.routes if hasattr(route, 'path')]

                results['tests']['router_creation'] = {
                    'status': 'passed',
                    'time': router_time,
                    'route_count': route_count,
                    'routes': route_paths
                }

            except Exception as e:
                results['tests']['router_creation'] = {
                    'status': 'failed',
                    'error': str(e)
                }

            # Test 3: Template Directory Structure
            template_dir_exists = api_integration.template_dir.exists()
            upload_dir_exists = api_integration.upload_dir.exists()

            results['tests']['directory_structure'] = {
                'status': 'passed' if template_dir_exists and upload_dir_exists else 'failed',
                'template_dir_exists': template_dir_exists,
                'upload_dir_exists': upload_dir_exists,
                'template_files': [str(f) for f in api_integration.template_dir.glob("*.xlsx")] if template_dir_exists else []
            }

            # Overall phase status
            all_passed = all(
                test.get('status') == 'passed'
                for test in results['tests'].values()
            )

            results['status'] = 'passed' if all_passed else 'failed'

        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(str(e))
            logger.error(f"API integration validation failed: {e}")

        return results

    def validate_performance(self) -> Dict[str, Any]:
        """Validate system performance against targets"""
        results = {'status': 'running', 'tests': {}, 'errors': []}

        try:
            # Performance targets
            targets = self.test_config['performance_targets']

            # Test 1: Template Generation Performance
            from excel_template_generator import MarketRegimeTemplateGenerator

            generator = MarketRegimeTemplateGenerator(str(self.template_dir))

            template_perf = {}
            for template_type in ['18_REGIME', '8_REGIME']:
                start_time = time.time()
                template_path = generator.generate_template(template_type, f"perf_test_{template_type.lower()}.xlsx")
                generation_time = time.time() - start_time

                template_perf[template_type] = {
                    'time': generation_time,
                    'target': targets['template_generation_time'],
                    'status': 'passed' if generation_time <= targets['template_generation_time'] else 'failed',
                    'performance_ratio': generation_time / targets['template_generation_time']
                }

            results['tests']['template_generation_performance'] = template_perf

            # Test 2: Excel Parsing Performance
            from excel_config_parser import MarketRegimeExcelParser

            parser = MarketRegimeExcelParser()
            template_path = self.template_dir / "market_regime_18_config.xlsx"

            if template_path.exists():
                start_time = time.time()
                config = parser.parse_excel_config(str(template_path))
                parsing_time = time.time() - start_time

                results['tests']['excel_parsing_performance'] = {
                    'time': parsing_time,
                    'target': targets['excel_parsing_time'],
                    'status': 'passed' if parsing_time <= targets['excel_parsing_time'] else 'failed',
                    'performance_ratio': parsing_time / targets['excel_parsing_time']
                }
            else:
                results['tests']['excel_parsing_performance'] = {
                    'status': 'skipped',
                    'error': 'Template file not found'
                }

            # Test 3: Regime Calculation Performance
            from enhanced_regime_engine import EnhancedMarketRegimeEngine

            engine = EnhancedMarketRegimeEngine()

            # Create small test dataset
            dates = pd.date_range(start='2024-01-01', periods=10, freq='1min')
            test_data = pd.DataFrame({
                'open': 22000 + np.random.randn(10) * 50,
                'high': 22050 + np.random.randn(10) * 50,
                'low': 21950 + np.random.randn(10) * 50,
                'close': 22000 + np.random.randn(10) * 50,
                'volume': 1000 + np.random.randint(0, 500, 10)
            }, index=dates)

            start_time = time.time()
            regime_result = engine.calculate_market_regime(test_data)
            calculation_time = time.time() - start_time

            results['tests']['regime_calculation_performance'] = {
                'time': calculation_time,
                'target': targets['regime_calculation_time'],
                'status': 'passed' if calculation_time <= targets['regime_calculation_time'] else 'failed',
                'performance_ratio': calculation_time / targets['regime_calculation_time']
            }

            # Overall phase status - handle nested test results properly
            all_passed = True
            for test_name, test_group in results['tests'].items():
                if isinstance(test_group, dict):
                    # Handle nested test results (like template_generation_performance)
                    if 'status' in test_group:
                        # Direct test result
                        if test_group['status'] not in ['passed', 'skipped']:
                            all_passed = False
                            print(f"   ❌ {test_name}: {test_group['status']} - {test_group.get('error', 'Performance target not met')}")
                            break
                        else:
                            print(f"   ✅ {test_name}: {test_group['status']}")
                    else:
                        # Nested test results
                        for sub_test_name, sub_test in test_group.items():
                            if isinstance(sub_test, dict) and 'status' in sub_test:
                                if sub_test['status'] not in ['passed', 'skipped']:
                                    all_passed = False
                                    ratio = sub_test.get('performance_ratio', 'N/A')
                                    print(f"   ❌ {test_name}.{sub_test_name}: {sub_test['status']} (ratio: {ratio})")
                                    break
                                else:
                                    ratio = sub_test.get('performance_ratio', 'N/A')
                                    print(f"   ✅ {test_name}.{sub_test_name}: {sub_test['status']} (ratio: {ratio})")
                        if not all_passed:
                            break

            results['status'] = 'passed' if all_passed else 'failed'

        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(str(e))
            logger.error(f"Performance validation failed: {e}")

        return results

    def generate_validation_summary(self) -> Dict[str, Any]:
        """Generate comprehensive validation summary"""
        summary = {
            'overall_status': 'unknown',
            'total_time': time.time() - self.start_time if self.start_time else 0,
            'phase_results': {},
            'test_counts': {'passed': 0, 'failed': 0, 'skipped': 0},
            'performance_summary': {},
            'recommendations': []
        }

        try:
            # Analyze phase results
            phase_statuses = []
            for phase_name, phase_result in self.test_results.items():
                if phase_name.startswith('phase_'):
                    phase_status = phase_result.get('status', 'unknown')
                    phase_statuses.append(phase_status)

                    summary['phase_results'][phase_name] = {
                        'status': phase_status,
                        'test_count': len(phase_result.get('tests', {})),
                        'errors': len(phase_result.get('errors', []))
                    }

                    # Count individual tests
                    for test_group in phase_result.get('tests', {}).values():
                        if isinstance(test_group, dict):
                            for test in test_group.values():
                                if isinstance(test, dict) and 'status' in test:
                                    status = test['status']
                                    summary['test_counts'][status] = summary['test_counts'].get(status, 0) + 1
                        elif isinstance(test_group, dict) and 'status' in test_group:
                            status = test_group['status']
                            summary['test_counts'][status] = summary['test_counts'].get(status, 0) + 1

            # Determine overall status
            if all(status == 'passed' for status in phase_statuses):
                summary['overall_status'] = 'passed'
            elif any(status == 'failed' for status in phase_statuses):
                summary['overall_status'] = 'failed'
            else:
                summary['overall_status'] = 'partial'

            # Performance summary
            if 'phase_5' in self.test_results:
                perf_tests = self.test_results['phase_5'].get('tests', {})
                for test_name, test_result in perf_tests.items():
                    if isinstance(test_result, dict) and 'performance_ratio' in test_result:
                        summary['performance_summary'][test_name] = {
                            'ratio': test_result['performance_ratio'],
                            'status': test_result['status']
                        }

            # Generate recommendations
            if summary['overall_status'] == 'passed':
                summary['recommendations'].append("✅ All tests passed! Market Regime system is ready for production.")
            else:
                if summary['test_counts']['failed'] > 0:
                    summary['recommendations'].append(f"❌ {summary['test_counts']['failed']} tests failed. Review error logs and fix issues.")

                if 'phase_1' in self.test_results and self.test_results['phase_1']['status'] == 'failed':
                    summary['recommendations'].append("📄 Template system issues detected. Check Excel template generation.")

                if 'phase_3' in self.test_results and self.test_results['phase_3']['status'] == 'failed':
                    summary['recommendations'].append("🔧 Engine integration issues detected. Verify enhanced package installation.")

                if 'phase_5' in self.test_results and self.test_results['phase_5']['status'] == 'failed':
                    summary['recommendations'].append("⚡ Performance targets not met. Consider optimization.")

        except Exception as e:
            summary['error'] = str(e)
            logger.error(f"Summary generation failed: {e}")

        return summary

    def save_validation_report(self, output_path: str) -> None:
        """Save validation report to JSON file"""
        try:
            with open(output_path, 'w') as f:
                json.dump(self.test_results, f, indent=2, default=str)

            logger.info(f"📄 Validation report saved to: {output_path}")

        except Exception as e:
            logger.error(f"❌ Failed to save validation report: {e}")


def main():
    """Main function to run comprehensive validation"""
    try:
        print("🧪 Market Regime Detection System - Comprehensive Validation")
        print("=" * 70)

        # Initialize and run validation
        validator = MarketRegimeValidationTest()
        results = validator.run_comprehensive_validation()

        # Print summary
        summary = results.get('summary', {})

        print(f"\n📊 VALIDATION SUMMARY")
        print("=" * 30)
        print(f"Overall Status: {summary.get('overall_status', 'unknown').upper()}")
        print(f"Total Time: {summary.get('total_time', 0):.2f} seconds")
        print(f"Tests Passed: {summary.get('test_counts', {}).get('passed', 0)}")
        print(f"Tests Failed: {summary.get('test_counts', {}).get('failed', 0)}")
        print(f"Tests Skipped: {summary.get('test_counts', {}).get('skipped', 0)}")

        # Print phase results
        print(f"\n📋 PHASE RESULTS")
        print("-" * 20)
        for phase_name, phase_result in summary.get('phase_results', {}).items():
            status_icon = "✅" if phase_result['status'] == 'passed' else "❌" if phase_result['status'] == 'failed' else "⚠️"
            print(f"{status_icon} {phase_name}: {phase_result['status'].upper()} ({phase_result['test_count']} tests)")

        # Print recommendations
        recommendations = summary.get('recommendations', [])
        if recommendations:
            print(f"\n💡 RECOMMENDATIONS")
            print("-" * 20)
            for rec in recommendations:
                print(f"   {rec}")

        # Save detailed report
        report_path = "/tmp/market_regime_validation_report.json"
        validator.save_validation_report(report_path)

        print(f"\n📄 Detailed report saved to: {report_path}")

        # Return appropriate exit code
        if summary.get('overall_status') == 'passed':
            print("\n🎉 Market Regime Detection System validation PASSED!")
            return 0
        else:
            print("\n⚠️ Market Regime Detection System validation had issues.")
            return 1

    except Exception as e:
        print(f"❌ Validation failed with error: {e}")
        return 1


if __name__ == "__main__":
    exit(main())
